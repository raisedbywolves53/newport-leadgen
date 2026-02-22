#!/usr/bin/env python3
"""
Build Newport GovCon Pro Forma Financial Model.

Generates a professional Excel workbook with 4 sheets:
  1. Assumptions  - All editable inputs (blue text, yellow highlights)
  2. Revenue Model - 60-month projection x 3 scenarios (all Excel formulas)
  3. Summary       - Executive view: Year 1-5 results + 5-year totals
  4. Platform Comparison - Free vs Optimal side-by-side

ALL calculations use Excel formulas, NOT Python math.
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
BLUE_FONT = Font(name="Calibri", size=11, color="0000FF")
BLUE_FONT_BOLD = Font(name="Calibri", size=11, color="0000FF", bold=True)
BLACK_FONT = Font(name="Calibri", size=11, color="000000")
BLACK_FONT_BOLD = Font(name="Calibri", size=11, color="000000", bold=True)
GREEN_FONT = Font(name="Calibri", size=11, color="008000")
GREEN_FONT_BOLD = Font(name="Calibri", size=11, color="008000", bold=True)
HEADER_FONT = Font(name="Calibri", size=12, color="000000", bold=True)
TITLE_FONT = Font(name="Calibri", size=14, color="000000", bold=True)
SECTION_FONT = Font(name="Calibri", size=11, color="000000", bold=True, italic=True)
NOTE_FONT = Font(name="Calibri", size=10, color="808080", italic=True)
COST_NOTE_FONT = Font(name="Calibri", size=10, color="8B4513", italic=True)

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FMT_CURRENCY = '$#,##0'
FMT_CURRENCY_LARGE = '$#,##0,,"M"'
FMT_PERCENT = '0.0%'
FMT_COUNT = '#,##0'
FMT_COUNT_1 = '0.0'

# Number of projection months
NUM_MONTHS = 60


def set_cell(ws, row, col, value, font=None, fill=None, fmt=None, alignment=None, border=None):
    """Helper to set a cell's value and styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def auto_fit_columns(ws, min_width=10, max_width=40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                if val_str.startswith("="):
                    val_str = val_str[:20]
                max_len = max(max_len, len(val_str))
        width = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def generate_month_labels(num_months=60, start_year=2026, start_month=3):
    """Generate calendar month labels: Mar 2026, Apr 2026, ..., Feb 2031."""
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    labels = []
    for i in range(num_months):
        m = (start_month - 1 + i) % 12
        y = start_year + (start_month - 1 + i) // 12
        labels.append(f"{month_names[m]} {y}")
    return labels


def load_market_data():
    """Try to load market_data.json for TAM context. Returns None values if unavailable."""
    market_data_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "market_data.json"
    )
    try:
        with open(market_data_path) as f:
            md = json.load(f)
        return {
            "tam_total": md.get("tam", {}).get("total_spending"),
            "tam_target": md.get("tam", {}).get("target_states_spending"),
            "fiscal_year": md.get("fiscal_year"),
        }
    except (FileNotFoundError, json.JSONDecodeError):
        return {"tam_total": None, "tam_target": None, "fiscal_year": None}


# ---------------------------------------------------------------------------
# Assumptions sheet row map (used by Revenue Model and Summary formulas)
# ---------------------------------------------------------------------------
# Business Assumptions
ROW_MARGIN = 3       # B3: Wholesale Gross Margin
ROW_CV_CON = 4       # B4: Avg Contract Value - Conservative
ROW_CV_MOD = 5       # B5: Avg Contract Value - Moderate
ROW_CV_AGG = 6       # B6: Avg Contract Value - Aggressive
ROW_DURATION = 7     # B7: Average Contract Duration (months)

# Win Rate Assumptions
ROW_WR_Y1_H1 = 11   # B11: Year 1 Win Rate (Months 1-6)
ROW_WR_Y1_H2 = 12   # B12: Year 1 Win Rate (Months 7-12)
ROW_WR_Y2 = 13       # B13: Year 2 Win Rate
ROW_WR_Y3 = 14       # B14: Year 3 Win Rate
ROW_WR_Y4 = 15       # B15: Year 4 Win Rate
ROW_WR_Y5 = 16       # B16: Year 5 Win Rate

# Bid Volume Assumptions
ROW_BID_CON = 19     # B19: Bids/Month - Conservative
ROW_BID_MOD = 20     # B20: Bids/Month - Moderate
ROW_BID_AGG = 21     # B21: Bids/Month - Aggressive
ROW_BID_MULT_Y3 = 22  # B22: Year 3 Bid Multiplier
ROW_BID_MULT_Y4 = 23  # B23: Year 4 Bid Multiplier
ROW_BID_MULT_Y5 = 24  # B24: Year 5 Bid Multiplier

# Platform Cost Assumptions
ROW_PLAT_FREE = 27   # B27: Free System Cost (annual)
ROW_PLAT_OPT = 31    # B31: Optimal System Total (annual)
ROW_CONSULTING = 32  # B32: Consulting Fee

# Ramp Assumptions
ROW_RAMP_BID = 35    # B35: Months to First Bid
ROW_RAMP_WIN = 36    # B36: Months to First Win


# ---------------------------------------------------------------------------
# Sheet 1: Assumptions
# ---------------------------------------------------------------------------
def build_assumptions(wb):
    ws = wb.active
    ws.title = "Assumptions"

    # Title
    set_cell(ws, 1, 1, "Newport GovCon Pro Forma Assumptions (5-Year)", font=TITLE_FONT)
    ws.merge_cells("A1:C1")

    # ---- Section: Newport Business Assumptions ----
    set_cell(ws, 2, 1, "Newport Business Assumptions", font=SECTION_FONT, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 2, 2, None, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 2, 3, None, fill=LIGHT_GRAY_FILL)

    labels_business = [
        (3, "Wholesale Gross Margin", 0.11, FMT_PERCENT, True,
         "Newport's estimated margin on government food supply"),
        (4, "Average Contract Value - Conservative", 50000, FMT_CURRENCY, True,
         "Micro-purchase range; FL avg $10K-$25K bracket = 54% of visible contracts"),
        (5, "Average Contract Value - Moderate", 75000, FMT_CURRENCY, True,
         "Simplified acquisition; FL avg $25K-$100K bracket = 33% of visible contracts"),
        (6, "Average Contract Value - Aggressive", 100000, FMT_CURRENCY, True,
         "Larger simplified; FL avg $100K-$350K bracket = 13% of visible contracts"),
        (7, "Average Contract Duration (months)", 12, FMT_COUNT, False, "Most small contracts are annual"),
        (8, "Revenue Recognition", "Monthly", None, False, "Spread evenly over contract duration"),
    ]
    for row, label, value, fmt, highlight, note in labels_business:
        set_cell(ws, row, 1, label, font=BLACK_FONT)
        set_cell(ws, row, 2, value, font=BLUE_FONT, fmt=fmt,
                 fill=YELLOW_FILL if highlight else None)
        if note:
            set_cell(ws, row, 3, note, font=NOTE_FONT)

    # ---- Section: Win Rate Assumptions (5-Year) ----
    set_cell(ws, 10, 1, "Win Rate Assumptions (5-Year)", font=SECTION_FONT, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 10, 2, None, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 10, 3, None, fill=LIGHT_GRAY_FILL)

    labels_winrate = [
        (ROW_WR_Y1_H1, "Year 1 Win Rate (Months 1-6)", 0.15, True,
         "New entrant — 93% sole source at DoD confirmed (FPDS); micro-purchases have 35-45% win rate"),
        (ROW_WR_Y1_H2, "Year 1 Win Rate (Months 7-12)", 0.25, True,
         "Building past performance; 83% of contracts are micro-purchases with low barriers"),
        (ROW_WR_Y2, "Year 2 Win Rate", 0.35, True,
         "Established past performance; FPDS shows 1.2 avg offers in target categories"),
        (ROW_WR_Y3, "Year 3 Win Rate", 0.40, True,
         "Strong past performance portfolio; research shows compounding to 35-40%"),
        (ROW_WR_Y4, "Year 4 Win Rate", 0.45, True,
         "Preferred vendor status; incumbent renewals beginning"),
        (ROW_WR_Y5, "Year 5 Win Rate", 0.50, True,
         "Incumbent advantage; research shows 40-50% at this stage"),
    ]
    for row, label, value, highlight, note in labels_winrate:
        set_cell(ws, row, 1, label, font=BLACK_FONT)
        set_cell(ws, row, 2, value, font=BLUE_FONT, fmt=FMT_PERCENT,
                 fill=YELLOW_FILL if highlight else None)
        if note:
            set_cell(ws, row, 3, note, font=NOTE_FONT)

    # ---- Section: Bid Volume Assumptions ----
    set_cell(ws, 18, 1, "Bid Volume Assumptions", font=SECTION_FONT, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 18, 2, None, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 18, 3, None, fill=LIGHT_GRAY_FILL)

    labels_bid = [
        (ROW_BID_CON, "Bids/Month - Conservative (Free system)", 1.5, FMT_COUNT_1, True, "~18/year"),
        (ROW_BID_MOD, "Bids/Month - Moderate (Optimal system)", 3, FMT_COUNT_1, True, "~36/year"),
        (ROW_BID_AGG, "Bids/Month - Aggressive (Optimal + dedicated effort)", 5, FMT_COUNT_1, True, "~60/year"),
        (ROW_BID_MULT_Y3, "Year 3 Bid Volume Multiplier", 1.25, FMT_COUNT_1, True, "25% increase — expanded capacity"),
        (ROW_BID_MULT_Y4, "Year 4 Bid Volume Multiplier", 1.50, FMT_COUNT_1, True, "50% increase — dedicated BD staff"),
        (ROW_BID_MULT_Y5, "Year 5 Bid Volume Multiplier", 1.75, FMT_COUNT_1, True, "75% increase — mature operation"),
    ]
    for row, label, value, fmt, highlight, note in labels_bid:
        set_cell(ws, row, 1, label, font=BLACK_FONT)
        set_cell(ws, row, 2, value, font=BLUE_FONT, fmt=fmt,
                 fill=YELLOW_FILL if highlight else None)
        if note:
            set_cell(ws, row, 3, note, font=NOTE_FONT)

    # ---- Section: Platform Cost Assumptions ----
    set_cell(ws, 26, 1, "Platform Cost Assumptions", font=SECTION_FONT, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 26, 2, None, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 26, 3, None, fill=LIGHT_GRAY_FILL)

    labels_platform = [
        (ROW_PLAT_FREE, "Free System Cost (annual)", 0, FMT_CURRENCY, False, "Built-in tools only"),
        (28, "CLEATUS (annual)", 3000, FMT_CURRENCY, True, "Mid-range estimate"),
        (29, "HigherGov (annual)", 3500, FMT_CURRENCY, True, "Mid-range estimate"),
        (30, "GovSpend (annual)", 6500, FMT_CURRENCY, True, "Mid-range estimate"),
        (ROW_PLAT_OPT, "Optimal System Total (annual)", "=SUM(B28:B30)", FMT_CURRENCY, False,
         "Formula: sum of platform costs"),
        (ROW_CONSULTING, "Consulting Fee - Monthly Retainer", None, FMT_CURRENCY, True,
         "Leave blank for Newport to discuss"),
    ]
    for row, label, value, fmt, highlight, note in labels_platform:
        set_cell(ws, row, 1, label, font=BLACK_FONT)
        font_to_use = BLACK_FONT if (isinstance(value, str) and str(value).startswith("=")) else BLUE_FONT
        set_cell(ws, row, 2, value, font=font_to_use, fmt=fmt,
                 fill=YELLOW_FILL if highlight else None)
        if note:
            set_cell(ws, row, 3, note, font=NOTE_FONT)

    # ---- Section: Ramp Assumptions ----
    set_cell(ws, 34, 1, "Ramp Assumptions", font=SECTION_FONT, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 34, 2, None, fill=LIGHT_GRAY_FILL)
    set_cell(ws, 34, 3, None, fill=LIGHT_GRAY_FILL)

    labels_ramp = [
        (ROW_RAMP_BID, "Months to First Bid", 2, FMT_COUNT, True,
         "SAM.gov registration + first opportunity cycle"),
        (ROW_RAMP_WIN, "Months to First Win", 4, FMT_COUNT, True,
         "Average award timeline after first submission"),
    ]
    for row, label, value, fmt, highlight, note in labels_ramp:
        set_cell(ws, row, 1, label, font=BLACK_FONT)
        set_cell(ws, row, 2, value, font=BLUE_FONT, fmt=fmt,
                 fill=YELLOW_FILL if highlight else None)
        if note:
            set_cell(ws, row, 3, note, font=NOTE_FONT)

    # ---- Section: Cost Notations (informational, not computed) ----
    set_cell(ws, 38, 1, "Cost Notations (Not Modeled — For Planning Reference)", font=SECTION_FONT, fill=LIGHT_ORANGE_FILL)
    set_cell(ws, 38, 2, None, fill=LIGHT_ORANGE_FILL)
    set_cell(ws, 38, 3, None, fill=LIGHT_ORANGE_FILL)

    cost_notes = [
        (39, "Bid Prep — Micro-Purchase ($3K-$15K)", "$100-$500/bid",
         "Pricing + quote format only; 83% of FL contracts are this size"),
        (40, "Bid Prep — Simplified ($15K-$350K)", "$2K-$5K/bid",
         "Technical writing + compliance; 14.4% of FL contracts"),
        (41, "Surety Bonding", "1-3% of contract value",
         "If required — most food supply contracts do not require bonds"),
        (42, "HACCP/Food Safety Compliance", "One-time + annual renewal",
         "Required for USDA commodity distribution programs"),
        (43, "Working Capital / Cash Flow Gap", "NET 30-90 day terms",
         "Federal NET 30; state/local NET 60-90. FL avg contract $55K"),
        (44, "Insurance (GL + Product Liability)", "Annual increase",
         "Government contracts may require higher coverage limits"),
        (45, "Dedicated BD Staff", "Year 3+ consideration",
         "Full-time BD hire when pipeline justifies; ~215-660 hrs/yr Year 1"),
    ]
    for row, label, estimate, note in cost_notes:
        set_cell(ws, row, 1, label, font=BLACK_FONT)
        set_cell(ws, row, 2, estimate, font=COST_NOTE_FONT)
        set_cell(ws, row, 3, note, font=NOTE_FONT)

    # ---- Section: TAM Context (Confirmed from Live API Data, Feb 2026) ----
    set_cell(ws, 46, 1, "TAM Context (Confirmed Live Data — Feb 2026)", font=SECTION_FONT, fill=LIGHT_GREEN_FILL)
    set_cell(ws, 46, 2, None, fill=LIGHT_GREEN_FILL)
    set_cell(ws, 46, 3, None, fill=LIGHT_GREEN_FILL)

    # Try market_data.json for dynamic TAM, fall back to confirmed research numbers
    md = load_market_data()
    tam_total = md["tam_total"]
    tam_target = md["tam_target"]

    set_cell(ws, 47, 1, "Total Federal Food Spending (FY2024)", font=BLACK_FONT)
    set_cell(ws, 47, 2, tam_total if tam_total else 7170000000,
             font=GREEN_FONT, fmt=FMT_CURRENCY)
    set_cell(ws, 47, 3, "PSC 89xx confirmed via USASpending + FPDS (HIGH confidence)", font=NOTE_FONT)

    set_cell(ws, 48, 1, "FL Food Contracts Under $350K (FY2024)", font=BLACK_FONT)
    set_cell(ws, 48, 2, 85000000, font=GREEN_FONT, fmt=FMT_CURRENCY)
    set_cell(ws, 48, 3, "39,685 awards — confirmed via USASpending (HIGH confidence)", font=NOTE_FONT)

    set_cell(ws, 49, 1, "FL Visible Contracts >$10K (FY2024)", font=BLACK_FONT)
    set_cell(ws, 49, 2, 6400000, font=GREEN_FONT, fmt=FMT_CURRENCY)
    set_cell(ws, 49, 3, "117 contracts — live USASpending report (HIGH confidence)", font=NOTE_FONT)

    set_cell(ws, 50, 1, "FL Micro-Purchases <$15K (Invisible to Free Tools)", font=BLACK_FONT)
    set_cell(ws, 50, 2, "~$8-15M/yr est.", font=COST_NOTE_FONT)
    set_cell(ws, 50, 3, "83% of contracts are micro-purchases — requires GovSpend ($6.5K/yr) to see", font=NOTE_FONT)

    set_cell(ws, 51, 1, "FL SLED Market (Schools, Corrections, etc.)", font=BLACK_FONT)
    set_cell(ws, 51, 2, "$600M-$1.2B/yr est.", font=COST_NOTE_FONT)
    set_cell(ws, 51, 3, "Not in FPDS/USASpending — requires HigherGov ($3.5K/yr) or state portals", font=NOTE_FONT)

    set_cell(ws, 52, 1, "Newport Serviceable Market (Year 1, Federal)", font=BLACK_FONT)
    set_cell(ws, 52, 2, "$2-5M biddable", font=COST_NOTE_FONT)
    set_cell(ws, 52, 3, "After filtering: categories, geography, past performance requirements", font=NOTE_FONT)

    # ---- Section: Confirmed Competition Data ----
    set_cell(ws, 54, 1, "Confirmed Competition Data (Live FPDS, Feb 2026)", font=SECTION_FONT, fill=LIGHT_GREEN_FILL)
    set_cell(ws, 54, 2, None, fill=LIGHT_GREEN_FILL)
    set_cell(ws, 54, 3, None, fill=LIGHT_GREEN_FILL)

    competition_data = [
        (55, "#1 FL Food Buyer: DOJ/Bureau of Prisons", "$3.7M / 71 contracts",
         "Live USASpending — direct competitive target (Rainmaker doing $5M)"),
        (56, "#2 FL Food Buyer: Dept of Defense", "$2.3M / 43 contracts",
         "MacDill, Homestead, NAS Jax, Patrick SFB — local base procurement"),
        (57, "NAICS 424490 (Other Grocery) @ DoD", "93% sole source",
         "117 awards, 1.2 avg offers — golden target, almost no competition"),
        (58, "FPDS Overall", "537 awards, 32 combos",
         "15 of 32 NAICS/agency combos = LOW competition (easy entry)"),
        (59, "FEMA Direct Food Procurement", "Narrow & concentrated",
         "11 contracts, 2 vendors = 97%. Play = disaster registry + micro-purchases"),
        (60, "Top FL Competitors", "$1-5M range",
         "Oakes Farms $26M, US Foods $24M, then small vendors $1-5M"),
    ]
    for row, label, value, note in competition_data:
        set_cell(ws, row, 1, label, font=BLACK_FONT)
        set_cell(ws, row, 2, value, font=GREEN_FONT)
        set_cell(ws, row, 3, note, font=NOTE_FONT)

    # Column widths
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 58

    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Revenue Model (60 months x 3 scenarios)
# ---------------------------------------------------------------------------
def build_revenue_model(wb):
    ws = wb.create_sheet("Revenue Model")

    months_label = generate_month_labels(NUM_MONTHS)

    # -- Row 1: Title + month headers --
    set_cell(ws, 1, 1, "Revenue Model - 60 Month Projection (5-Year GovCon Plan)", font=TITLE_FONT)
    for m in range(NUM_MONTHS):
        col = m + 2
        set_cell(ws, 1, col, None, fill=LIGHT_GRAY_FILL)

    # -- Row 2: Month numbers 1-60 --
    set_cell(ws, 2, 1, "Month #", font=HEADER_FONT, fill=LIGHT_GRAY_FILL)
    for m in range(NUM_MONTHS):
        col = m + 2
        set_cell(ws, 2, col, m + 1, font=BLACK_FONT_BOLD, fill=LIGHT_GRAY_FILL,
                 alignment=Alignment(horizontal="center"))

    # -- Row 3: Calendar month labels --
    set_cell(ws, 3, 1, "Calendar Month", font=HEADER_FONT, fill=LIGHT_GRAY_FILL)
    for m in range(NUM_MONTHS):
        col = m + 2
        set_cell(ws, 3, col, months_label[m], font=BLACK_FONT, fill=LIGHT_GRAY_FILL,
                 alignment=Alignment(horizontal="center"))

    # -------------------------------------------------------------------
    # Scenario definitions
    # Layout: Conservative rows 5-16, Moderate rows 18-29, Aggressive rows 31-42
    # -------------------------------------------------------------------
    scenarios = [
        ("CONSERVATIVE SCENARIO",  4,  5, f"B{ROW_BID_CON}", f"B{ROW_CV_CON}", f"=Assumptions!$B${ROW_PLAT_FREE}/12"),
        ("MODERATE SCENARIO",     17, 18, f"B{ROW_BID_MOD}", f"B{ROW_CV_MOD}", f"=Assumptions!$B${ROW_PLAT_OPT}/12"),
        ("AGGRESSIVE SCENARIO",   30, 31, f"B{ROW_BID_AGG}", f"B{ROW_CV_AGG}", f"=Assumptions!$B${ROW_PLAT_OPT}/12"),
    ]

    metric_labels = [
        "Bids Submitted",
        "Win Rate",
        "New Contracts Won",
        "Cumulative Contracts Won",
        "Active Contracts",
        "Monthly Revenue",
        "Cumulative Revenue",
        "Monthly Gross Profit",
        "Cumulative Gross Profit",
        "Monthly Platform Cost",
        "Monthly Net Contribution",
        "Cumulative Net Contribution",
    ]

    metric_fmts = [
        FMT_COUNT_1,   # Bids Submitted
        FMT_PERCENT,   # Win Rate
        FMT_COUNT,     # New Contracts Won
        FMT_COUNT,     # Cumulative Contracts Won
        FMT_COUNT,     # Active Contracts
        FMT_CURRENCY,  # Monthly Revenue
        FMT_CURRENCY,  # Cumulative Revenue
        FMT_CURRENCY,  # Monthly Gross Profit
        FMT_CURRENCY,  # Cumulative Gross Profit
        FMT_CURRENCY,  # Monthly Platform Cost
        FMT_CURRENCY,  # Monthly Net Contribution
        FMT_CURRENCY,  # Cumulative Net Contribution
    ]

    for scenario_title, hdr_row, sr, bid_ref, cv_ref, plat_formula in scenarios:
        # Scenario header row
        set_cell(ws, hdr_row, 1, scenario_title, font=HEADER_FONT, fill=LIGHT_BLUE_FILL)
        for m in range(NUM_MONTHS):
            set_cell(ws, hdr_row, m + 2, None, fill=LIGHT_BLUE_FILL)

        # Metric labels
        for i, label in enumerate(metric_labels):
            row = sr + i
            set_cell(ws, row, 1, label, font=BLACK_FONT_BOLD)

        # Metric row positions
        r_bids    = sr + 0
        r_wr      = sr + 1
        r_new     = sr + 2
        r_cum     = sr + 3
        r_active  = sr + 4
        r_rev     = sr + 5
        r_cumrev  = sr + 6
        r_gp      = sr + 7
        r_cumgp   = sr + 8
        r_plat    = sr + 9
        r_net     = sr + 10
        r_cumnet  = sr + 11

        for m in range(NUM_MONTHS):
            col = m + 2
            col_letter = get_column_letter(col)

            # ----- Bids Submitted -----
            # Base bid volume * multiplier for Years 3-5
            # =IF(month<first_bid, 0, base_bids * IF(month>48,Y5_mult, IF(month>36,Y4_mult, IF(month>24,Y3_mult, 1))))
            f_bids = (
                f'=IF({col_letter}$2<Assumptions!$B${ROW_RAMP_BID},0,'
                f'Assumptions!${bid_ref[0]}${bid_ref[1:]}'
                f'*IF({col_letter}$2>48,Assumptions!$B${ROW_BID_MULT_Y5},'
                f'IF({col_letter}$2>36,Assumptions!$B${ROW_BID_MULT_Y4},'
                f'IF({col_letter}$2>24,Assumptions!$B${ROW_BID_MULT_Y3},1))))'
            )
            set_cell(ws, r_bids, col, f_bids, font=BLACK_FONT, fmt=metric_fmts[0])

            # ----- Win Rate -----
            # 6-tier: Y1H1, Y1H2, Y2, Y3, Y4, Y5
            f_wr = (
                f'=IF({col_letter}$2<Assumptions!$B${ROW_RAMP_WIN},0,'
                f'IF({col_letter}$2<=6,Assumptions!$B${ROW_WR_Y1_H1},'
                f'IF({col_letter}$2<=12,Assumptions!$B${ROW_WR_Y1_H2},'
                f'IF({col_letter}$2<=24,Assumptions!$B${ROW_WR_Y2},'
                f'IF({col_letter}$2<=36,Assumptions!$B${ROW_WR_Y3},'
                f'IF({col_letter}$2<=48,Assumptions!$B${ROW_WR_Y4},'
                f'Assumptions!$B${ROW_WR_Y5}))))))'
            )
            set_cell(ws, r_wr, col, f_wr, font=GREEN_FONT, fmt=metric_fmts[1])

            # ----- New Contracts Won -----
            f_new = f'=ROUND({col_letter}{r_bids}*{col_letter}{r_wr},0)'
            set_cell(ws, r_new, col, f_new, font=BLACK_FONT, fmt=metric_fmts[2])

            # ----- Cumulative Contracts Won -----
            if m == 0:
                f_cum = f'={col_letter}{r_new}'
            else:
                prev_col = get_column_letter(col - 1)
                f_cum = f'={prev_col}{r_cum}+{col_letter}{r_new}'
            set_cell(ws, r_cum, col, f_cum, font=BLACK_FONT, fmt=metric_fmts[3])

            # ----- Active Contracts -----
            f_active = (
                f'={col_letter}{r_cum}'
                f'-IF({col_letter}$2>Assumptions!$B${ROW_DURATION},'
                f'OFFSET({col_letter}{r_cum},0,-Assumptions!$B${ROW_DURATION}),0)'
            )
            set_cell(ws, r_active, col, f_active, font=BLACK_FONT, fmt=metric_fmts[4])

            # ----- Monthly Revenue -----
            f_rev = (
                f'={col_letter}{r_active}'
                f'*(Assumptions!${cv_ref[0]}${cv_ref[1:]}/Assumptions!$B${ROW_DURATION})'
            )
            set_cell(ws, r_rev, col, f_rev, font=GREEN_FONT, fmt=metric_fmts[5])

            # ----- Cumulative Revenue -----
            if m == 0:
                f_cumrev = f'={col_letter}{r_rev}'
            else:
                prev_col = get_column_letter(col - 1)
                f_cumrev = f'={prev_col}{r_cumrev}+{col_letter}{r_rev}'
            set_cell(ws, r_cumrev, col, f_cumrev, font=BLACK_FONT, fmt=metric_fmts[6])

            # ----- Monthly Gross Profit -----
            f_gp = f'={col_letter}{r_rev}*Assumptions!$B${ROW_MARGIN}'
            set_cell(ws, r_gp, col, f_gp, font=GREEN_FONT, fmt=metric_fmts[7])

            # ----- Cumulative Gross Profit -----
            if m == 0:
                f_cumgp = f'={col_letter}{r_gp}'
            else:
                prev_col = get_column_letter(col - 1)
                f_cumgp = f'={prev_col}{r_cumgp}+{col_letter}{r_gp}'
            set_cell(ws, r_cumgp, col, f_cumgp, font=BLACK_FONT, fmt=metric_fmts[8])

            # ----- Monthly Platform Cost -----
            set_cell(ws, r_plat, col, plat_formula, font=GREEN_FONT, fmt=metric_fmts[9])

            # ----- Monthly Net Contribution -----
            f_net = f'={col_letter}{r_gp}-{col_letter}{r_plat}'
            set_cell(ws, r_net, col, f_net, font=BLACK_FONT, fmt=metric_fmts[10])

            # ----- Cumulative Net Contribution -----
            if m == 0:
                f_cumnet = f'={col_letter}{r_net}'
            else:
                prev_col = get_column_letter(col - 1)
                f_cumnet = f'={prev_col}{r_cumnet}+{col_letter}{r_net}'
            set_cell(ws, r_cumnet, col, f_cumnet, font=BLACK_FONT, fmt=metric_fmts[11])

    # Freeze panes
    ws.freeze_panes = "B4"

    # Column widths
    ws.column_dimensions["A"].width = 30
    for m in range(NUM_MONTHS):
        ws.column_dimensions[get_column_letter(m + 2)].width = 14

    return ws


# ---------------------------------------------------------------------------
# Sheet 3: Summary (5-Year)
# ---------------------------------------------------------------------------
def build_summary(wb):
    ws = wb.create_sheet("Summary")

    # Title
    set_cell(ws, 1, 1, "Newport GovCon 5-Year Financial Summary", font=TITLE_FONT)
    ws.merge_cells("A1:D1")

    # Headers
    headers = ["Metric", "Conservative", "Moderate", "Aggressive"]
    for i, h in enumerate(headers):
        set_cell(ws, 3, i + 1, h, font=HEADER_FONT, fill=LIGHT_GRAY_FILL,
                 alignment=Alignment(horizontal="center"))

    # Scenario first-metric rows on Revenue Model (unchanged layout)
    srs = [5, 18, 31]

    # Year definitions: (label, start_month, end_month, start_col_letter, end_col_letter, end_col_letter_for_point)
    # Columns: month N is in column N+1 (B=month1, C=month2, ...)
    # Year 1: months 1-12  → cols B(2) to M(13)
    # Year 2: months 13-24 → cols N(14) to Y(25)
    # Year 3: months 25-36 → cols Z(26) to AK(37)
    # Year 4: months 37-48 → cols AL(38) to AW(49)
    # Year 5: months 49-60 → cols AX(50) to BI(61)
    years = [
        ("YEAR 1 RESULTS (Months 1-12)",  2,  13),   # col B to M
        ("YEAR 2 RESULTS (Months 13-24)", 14, 25),    # col N to Y
        ("YEAR 3 RESULTS (Months 25-36)", 26, 37),    # col Z to AK
        ("YEAR 4 RESULTS (Months 37-48)", 38, 49),    # col AL to AW
        ("YEAR 5 RESULTS (Months 49-60)", 50, 61),    # col AX to BI
    ]

    row_cursor = 5

    for year_idx, (year_label, start_col, end_col) in enumerate(years):
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)

        # Section header
        set_cell(ws, row_cursor, 1, year_label, font=HEADER_FONT, fill=LIGHT_BLUE_FILL)
        for c in range(2, 5):
            set_cell(ws, row_cursor, c, None, fill=LIGHT_BLUE_FILL)
        row_cursor += 1

        if year_idx == 0:
            # Year 1: special metrics (cumulative at month 12 for some)
            year_metrics = [
                ("Total Bids Submitted", 0, FMT_COUNT_1, "sum"),
                ("Contracts Won", 3, FMT_COUNT, "point"),
                ("Total Revenue", 6, FMT_CURRENCY, "point"),
                ("Gross Profit", 8, FMT_CURRENCY, "point"),
                ("Platform Investment", 9, FMT_CURRENCY, "sum"),
                ("Net Contribution", 11, FMT_CURRENCY, "point"),
            ]
            for label, offset, fmt, agg in year_metrics:
                fill = ALT_ROW_FILL if (row_cursor % 2 == 0) else WHITE_FILL
                set_cell(ws, row_cursor, 1, label, font=BLACK_FONT, fill=fill)
                for s_idx, sr in enumerate(srs):
                    metric_row = sr + offset
                    if agg == "sum":
                        formula = f"=SUM('Revenue Model'!{start_letter}{metric_row}:{end_letter}{metric_row})"
                    else:
                        formula = f"='Revenue Model'!{end_letter}{metric_row}"
                    set_cell(ws, row_cursor, s_idx + 2, formula, font=GREEN_FONT, fmt=fmt, fill=fill)
                row_cursor += 1
        else:
            # Years 2-5: sum-based metrics (incremental for that year)
            year_metrics = [
                ("Total Bids Submitted", 0, FMT_COUNT_1),
                ("Contracts Won", 2, FMT_COUNT),
                ("Total Revenue", 5, FMT_CURRENCY),
                ("Gross Profit", 7, FMT_CURRENCY),
                ("Platform Investment", 9, FMT_CURRENCY),
                ("Net Contribution", 10, FMT_CURRENCY),
            ]
            for label, offset, fmt in year_metrics:
                fill = ALT_ROW_FILL if (row_cursor % 2 == 0) else WHITE_FILL
                set_cell(ws, row_cursor, 1, label, font=BLACK_FONT, fill=fill)
                for s_idx, sr in enumerate(srs):
                    metric_row = sr + offset
                    formula = f"=SUM('Revenue Model'!{start_letter}{metric_row}:{end_letter}{metric_row})"
                    set_cell(ws, row_cursor, s_idx + 2, formula, font=GREEN_FONT, fmt=fmt, fill=fill)
                row_cursor += 1

        # ROI row
        fill = ALT_ROW_FILL if (row_cursor % 2 == 0) else WHITE_FILL
        set_cell(ws, row_cursor, 1, "ROI", font=BLACK_FONT_BOLD, fill=fill)
        for s_idx in range(3):
            col_letter = get_column_letter(s_idx + 2)
            net_row = row_cursor - 1
            plat_row = row_cursor - 2
            formula = f'=IF({col_letter}{plat_row}=0,"N/A (Free)",{col_letter}{net_row}/{col_letter}{plat_row})'
            set_cell(ws, row_cursor, s_idx + 2, formula, font=BLACK_FONT_BOLD, fmt=FMT_PERCENT, fill=fill)
        row_cursor += 2

    # -------------------------------------------------------------------
    # Section: 5-YEAR TOTALS
    # -------------------------------------------------------------------
    last_col_letter = get_column_letter(NUM_MONTHS + 1)  # BI for month 60

    set_cell(ws, row_cursor, 1, "5-YEAR TOTALS", font=HEADER_FONT, fill=LIGHT_BLUE_FILL)
    for c in range(2, 5):
        set_cell(ws, row_cursor, c, None, fill=LIGHT_BLUE_FILL)
    row_cursor += 1

    total_metrics = [
        ("Total Bids Submitted", 0, FMT_COUNT_1, "sum"),
        ("Total Contracts Won", 3, FMT_COUNT, "point"),
        ("Total Revenue", 6, FMT_CURRENCY, "point"),
        ("Total Gross Profit", 8, FMT_CURRENCY, "point"),
        ("Total Platform Investment", 9, FMT_CURRENCY, "sum"),
        ("Total Net Contribution", 11, FMT_CURRENCY, "point"),
    ]

    for label, offset, fmt, agg in total_metrics:
        fill = ALT_ROW_FILL if (row_cursor % 2 == 0) else WHITE_FILL
        set_cell(ws, row_cursor, 1, label, font=BLACK_FONT, fill=fill)
        for s_idx, sr in enumerate(srs):
            metric_row = sr + offset
            if agg == "sum":
                formula = f"=SUM('Revenue Model'!B{metric_row}:{last_col_letter}{metric_row})"
            else:
                formula = f"='Revenue Model'!{last_col_letter}{metric_row}"
            set_cell(ws, row_cursor, s_idx + 2, formula, font=GREEN_FONT, fmt=fmt, fill=fill)
        row_cursor += 1

    # 5-Year ROI
    fill = ALT_ROW_FILL if (row_cursor % 2 == 0) else WHITE_FILL
    set_cell(ws, row_cursor, 1, "5-Year ROI", font=BLACK_FONT_BOLD, fill=fill)
    for s_idx in range(3):
        col_letter = get_column_letter(s_idx + 2)
        net_row = row_cursor - 1
        plat_row = row_cursor - 2
        formula = f'=IF({col_letter}{plat_row}=0,"N/A (Free)",{col_letter}{net_row}/{col_letter}{plat_row})'
        set_cell(ws, row_cursor, s_idx + 2, formula, font=BLACK_FONT_BOLD, fmt=FMT_PERCENT, fill=fill)
    row_cursor += 1

    # Breakeven Month (across all 60 months)
    fill = ALT_ROW_FILL if (row_cursor % 2 == 0) else WHITE_FILL
    set_cell(ws, row_cursor, 1, "Breakeven Month", font=BLACK_FONT_BOLD, fill=fill)
    for s_idx, sr in enumerate(srs):
        cum_net_row = sr + 11
        formula = (
            f'=IFERROR(MATCH(TRUE,INDEX(\'Revenue Model\'!B{cum_net_row}:{last_col_letter}{cum_net_row}>0,),0),"Not in 60mo")'
        )
        set_cell(ws, row_cursor, s_idx + 2, formula, font=BLACK_FONT_BOLD, fmt=FMT_COUNT, fill=fill)

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    return ws


# ---------------------------------------------------------------------------
# Sheet 4: Platform Comparison
# ---------------------------------------------------------------------------
def build_platform_comparison(wb):
    ws = wb.create_sheet("Platform Comparison")

    # Title
    set_cell(ws, 1, 1, "Platform Comparison: Free vs. Optimal Investment", font=TITLE_FONT)
    ws.merge_cells("A1:C1")

    # Headers
    headers = ["Component", "Free", "Optimal"]
    for i, h in enumerate(headers):
        set_cell(ws, 3, i + 1, h, font=HEADER_FONT, fill=LIGHT_GRAY_FILL,
                 alignment=Alignment(horizontal="center"))

    # Data rows
    comparison_data = [
        ("Federal Monitoring", 'SAM.gov API - $0', '+ CLEATUS - $3,000/yr', False),
        ("State/Local (FL $600M-$1.2B)", 'Manual - $0', 'HigherGov - $3,500/yr', False),
        ("Micro-Purchases (83% of FL contracts)", 'Invisible - $0', 'GovSpend - $6,500/yr', False),
        ("Competitive Intelligence", 'FPDS + USASpending - $0', 'Same - $0', False),
        ("Proposal Writing", 'Claude + Templates - $0', 'CLEATUS AI - included', False),
        ("Pipeline Tracking", 'Google Sheets - $0', 'Same - $0', False),
    ]

    row_cursor = 4
    for label, free_val, optimal_val, is_total in comparison_data:
        fill = ALT_ROW_FILL if (row_cursor % 2 == 0) else WHITE_FILL
        set_cell(ws, row_cursor, 1, label, font=BLACK_FONT, fill=fill)
        set_cell(ws, row_cursor, 2, free_val, font=BLACK_FONT, fill=fill,
                 alignment=Alignment(horizontal="center"))
        set_cell(ws, row_cursor, 3, optimal_val, font=BLACK_FONT, fill=fill,
                 alignment=Alignment(horizontal="center"))
        row_cursor += 1

    # Separator
    row_cursor += 1

    # Totals row
    set_cell(ws, row_cursor, 1, "Total Annual Cost", font=BLACK_FONT_BOLD, fill=LIGHT_BLUE_FILL)
    set_cell(ws, row_cursor, 2, f"=Assumptions!B{ROW_PLAT_FREE}", font=GREEN_FONT_BOLD, fmt=FMT_CURRENCY,
             fill=LIGHT_BLUE_FILL, alignment=Alignment(horizontal="center"))
    set_cell(ws, row_cursor, 3, f"=Assumptions!B{ROW_PLAT_OPT}", font=GREEN_FONT_BOLD, fmt=FMT_CURRENCY,
             fill=LIGHT_BLUE_FILL, alignment=Alignment(horizontal="center"))
    row_cursor += 1

    # Additional comparison rows
    set_cell(ws, row_cursor, 1, "Market Coverage", font=BLACK_FONT_BOLD)
    set_cell(ws, row_cursor, 2, "~40-50%", font=BLACK_FONT,
             alignment=Alignment(horizontal="center"))
    set_cell(ws, row_cursor, 3, "~90%+", font=BLACK_FONT,
             alignment=Alignment(horizontal="center"))
    row_cursor += 1

    set_cell(ws, row_cursor, 1, "Bid Capacity/Year", font=BLACK_FONT_BOLD)
    set_cell(ws, row_cursor, 2, "12-20", font=BLACK_FONT,
             alignment=Alignment(horizontal="center"))
    set_cell(ws, row_cursor, 3, "40-60", font=BLACK_FONT,
             alignment=Alignment(horizontal="center"))
    row_cursor += 1

    set_cell(ws, row_cursor, 1, "Estimated Win Rate Uplift", font=BLACK_FONT_BOLD)
    set_cell(ws, row_cursor, 2, "Baseline", font=BLACK_FONT,
             alignment=Alignment(horizontal="center"))
    set_cell(ws, row_cursor, 3, "+30-50% more bids = more wins", font=BLACK_FONT,
             alignment=Alignment(horizontal="center"))
    row_cursor += 2

    # Monthly cost breakdown
    set_cell(ws, row_cursor, 1, "Monthly Cost Breakdown", font=HEADER_FONT, fill=LIGHT_GRAY_FILL)
    set_cell(ws, row_cursor, 2, None, fill=LIGHT_GRAY_FILL)
    set_cell(ws, row_cursor, 3, None, fill=LIGHT_GRAY_FILL)
    row_cursor += 1

    set_cell(ws, row_cursor, 1, "Monthly Platform Cost", font=BLACK_FONT)
    set_cell(ws, row_cursor, 2, f"=Assumptions!B{ROW_PLAT_FREE}/12", font=GREEN_FONT, fmt=FMT_CURRENCY,
             alignment=Alignment(horizontal="center"))
    set_cell(ws, row_cursor, 3, f"=Assumptions!B{ROW_PLAT_OPT}/12", font=GREEN_FONT, fmt=FMT_CURRENCY,
             alignment=Alignment(horizontal="center"))
    row_cursor += 1

    set_cell(ws, row_cursor, 1, "Consulting Retainer", font=BLACK_FONT)
    set_cell(ws, row_cursor, 2, "N/A", font=BLACK_FONT,
             alignment=Alignment(horizontal="center"))
    set_cell(ws, row_cursor, 3, f'=IF(Assumptions!B{ROW_CONSULTING}="","TBD",Assumptions!B{ROW_CONSULTING})',
             font=GREEN_FONT, fmt=FMT_CURRENCY, alignment=Alignment(horizontal="center"))

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35

    return ws


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------
def main():
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "newport-govcon-proforma.xlsx")

    print("=" * 60)
    print("Newport GovCon Pro Forma Financial Model Builder (5-Year)")
    print("=" * 60)

    wb = Workbook()

    print("[1/4] Building Assumptions sheet (5-year win rates, cost notations, TAM)...")
    build_assumptions(wb)

    print("[2/4] Building Revenue Model sheet (60 months x 3 scenarios)...")
    build_revenue_model(wb)

    print("[3/4] Building Summary sheet (Year 1-5 + 5-year totals)...")
    build_summary(wb)

    print("[4/4] Building Platform Comparison sheet...")
    build_platform_comparison(wb)

    # Save
    wb.save(output_path)
    print()
    print(f"Saved: {output_path}")
    print(f"File size: {os.path.getsize(output_path):,} bytes")
    print()

    # Verification
    print("Verification:")
    print(f"  Sheets: {wb.sheetnames}")
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"  {ws_name}: {ws.max_row} rows x {ws.max_column} cols")

    print()
    print("All calculations use Excel formulas (=SUM, =IF, =ROUND, =OFFSET, etc.)")
    print("Blue text = editable inputs | Black text = formulas | Green text = cross-sheet refs")
    print("Yellow background = key assumptions to review")
    print("Orange background = cost notations (informational)")
    print("Green background = TAM context (from market_data.json)")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
