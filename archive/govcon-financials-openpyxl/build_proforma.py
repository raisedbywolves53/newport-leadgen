#!/usr/bin/env python3
"""
Build Newport GovCon Financial Model (v7).

WIP v7 REWRITE — NOT YET THE CANONICAL MODEL.
The canonical deliverable is data/Newport_GovCon_Financial_Model_v7.xlsx
(built in Claude Desktop, 176 formulas, zero errors, 7 charts).
This script is a programmatic rewrite targeting v7 parity.
Status: ~95% complete — all 5 sheet builders implemented, charts included.
Deferred to Phase 5 for final validation and testing against the .xlsx.

Generates a 5-sheet Excel workbook matching the v7 planning tool:
  1. Inputs         - Newport assumptions, win rates, costs, renewals, investment
  2. Two Routes     - Free vs Paid route comparison
  3. 5-Year Model   - Bids, wins, contracts, revenue, owner earnings + 5 charts
  4. Market Analysis - TAM waterfall, contract examples, competition + 2 charts
  5. Key Questions   - 10 critical decisions

ALL calculations use Excel formulas, NOT Python math.

Usage:
    python govcon/deliverables/financials/build_proforma.py
    python govcon/deliverables/financials/build_proforma.py --output data/final/model.xlsx
"""

import argparse
import json
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Design system (matches v7 exactly)
# ---------------------------------------------------------------------------
# Colors
DARK_NAVY = "1B2A4A"
MEDIUM_BLUE = "2E5090"
PURE_BLUE = "0000FF"
GREEN = "548235"
ORANGE = "ED7D31"
RED = "C00000"
GRAY = "808080"
DARK_GRAY = "404040"
WHITE = "FFFFFF"
LIGHT_GRAY_HEX = "F2F2F2"
PALE_YELLOW = "FFF2CC"

# Fonts
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=11, color=WHITE)
BANNER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=11, color=DARK_GRAY)
LABEL_BOLD = Font(name="Calibri", bold=True, size=11, color=DARK_GRAY)
INPUT_FONT = Font(name="Calibri", size=11, color=PURE_BLUE)
INPUT_BOLD = Font(name="Calibri", bold=True, size=11, color=PURE_BLUE)
NOTE_FONT = Font(name="Calibri", size=10, color=GRAY, italic=True)
GREEN_FONT = Font(name="Calibri", size=11, color=GREEN)
GREEN_BOLD = Font(name="Calibri", bold=True, size=11, color=GREEN)
ORANGE_FONT = Font(name="Calibri", size=11, color=ORANGE)
ORANGE_BOLD = Font(name="Calibri", bold=True, size=11, color=ORANGE)
RED_FONT = Font(name="Calibri", size=11, color=RED)
RED_BOLD = Font(name="Calibri", bold=True, size=11, color=RED)
BLUE_FONT = Font(name="Calibri", size=11, color=MEDIUM_BLUE)
BLUE_BOLD = Font(name="Calibri", bold=True, size=11, color=MEDIUM_BLUE)
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color=DARK_GRAY)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
FORMULA_FONT = Font(name="Calibri", size=11, color="000000")

# Fills
NAVY_FILL = PatternFill("solid", fgColor=DARK_NAVY)
BLUE_FILL = PatternFill("solid", fgColor=MEDIUM_BLUE)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)
LGRAY_FILL = PatternFill("solid", fgColor=LIGHT_GRAY_HEX)
YELLOW_FILL = PatternFill("solid", fgColor=PALE_YELLOW)

# Borders
_THIN = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Number formats
FMT_CUR = '$#,##0;"($"#,##0);-'
FMT_PCT = "0.0%"
FMT_NUM = "#,##0"
FMT_CUR_M = '$#,##0,,"M"'

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sc(ws, r, c, val, font=None, fill=None, fmt=None, align=None):
    """Set cell value + style."""
    cell = ws.cell(row=r, column=c, value=val)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    cell.border = THIN_BORDER
    return cell


def banner(ws, row, col_start, col_end, text, fill=None):
    """Write a merged banner row."""
    f = fill or BLUE_FILL
    sc(ws, row, col_start, text, font=BANNER_FONT, fill=f, align=CENTER)
    for c in range(col_start + 1, col_end + 1):
        sc(ws, row, c, None, fill=f)
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end,
    )


def title_rows(ws, col_end):
    """Write standard title + subtitle in rows 2-3."""
    now = datetime.now()
    sc(ws, 2, 2, "Newport Wholesalers \u2014 GovCon Financial Model",
       font=TITLE_FONT, fill=NAVY_FILL, align=CENTER)
    for c in range(3, col_end + 1):
        sc(ws, 2, c, None, fill=NAVY_FILL)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=col_end)
    ws.row_dimensions[2].height = 20

    sc(ws, 3, 2, f"5-Year Projection | {now.strftime('%B %Y')}",
       font=SUBTITLE_FONT, fill=NAVY_FILL, align=CENTER)
    for c in range(3, col_end + 1):
        sc(ws, 3, c, None, fill=NAVY_FILL)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=col_end)


def zebra(row_num):
    """Return alternating fill for zebra striping."""
    return LGRAY_FILL if row_num % 2 == 0 else WHITE_FILL


def load_market_data():
    """Load market_data.json if available."""
    md_path = Path(__file__).resolve().parent.parent / "market_data.json"
    try:
        with open(md_path) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


# ---------------------------------------------------------------------------
# Sheet 1: Inputs
# ---------------------------------------------------------------------------
def build_inputs(wb):
    ws = wb.active
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    for c_letter in "CDEFG":
        ws.column_dimensions[c_letter].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 32

    title_rows(ws, 8)

    # ── NEWPORT INPUTS ─────────────────────────────────────────
    banner(ws, 5, 2, 9, "NEWPORT INPUTS")

    inputs = [
        (7,  "Current Annual Revenue", "Scale of gov entry relative to core business"),
        (8,  "Current Gross Margin (%)", "Baseline for government pricing strategy"),
        (9,  "Delivery Radius (miles from FL warehouse)", "FL only = $85M TAM. SE region = $179M TAM"),
        (10, "Refrigerated/Temp-Controlled Delivery? (Y/N)", "Opens or closes entire product categories"),
        (11, "# of Confectionery/Snack/Nut SKUs", "More SKUs = eligible for more bid opportunities"),
        (12, "Food Safety Certifications? (SQF/GFSI/FDA)", "$0 if certified. Up to $23.5K if not"),
        (13, "SDB/HUBZone/8(a) Eligible? (Y/N)", "Set-aside win rates 28% vs 8% open competition"),
        (14, "Min Profitable Order Size ($)", "Must be below $5,200 avg micro-purchase threshold"),
        (15, "Any Prior Gov Contract Experience?", "Past performance is the #1 evaluation factor"),
    ]
    for r, label, note in inputs:
        sc(ws, r, 2, label, font=LABEL_FONT)
        # Merge C:H for user input area (leave empty)
        for c in range(3, 9):
            sc(ws, r, c, None)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        sc(ws, r, 9, note, font=NOTE_FONT, align=LEFT_WRAP)

    # ── WIN RATE ASSUMPTIONS ───────────────────────────────────
    banner(ws, 17, 2, 9, "WIN RATE ASSUMPTIONS")

    # Header row
    yr_headers = ["Tier", "Yr 1", "Yr 2", "Yr 3", "Yr 4", "Yr 5", "", "Notes"]
    for i, h in enumerate(yr_headers):
        sc(ws, 18, 2 + i, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    # Win rate data: (row, label, yr1-yr5 values, note)
    wr_data = [
        (19, "Micro (<$15K)", [0.15, 0.20, 0.22, 0.25, 0.25],
         "GPC credit card - low barrier, high volume"),
        (20, "Simplified ($15K-$250K)", [0.05, 0.10, 0.14, 0.17, 0.20],
         "Starts Q3 Y1 with past perf refs"),
        (21, "Set-Aside (SDB/HUBZone)", [0.15, 0.22, 0.28, 0.32, 0.35],
         "IF certified. Day 1 eligible"),
        (22, "SLED (State/Local/Ed)", [0.08, 0.12, 0.16, 0.20, 0.22],
         "Relationship-driven, FL focus"),
        (23, "Sealed Bid ($250K+)", [0.00, 0.03, 0.06, 0.09, 0.12],
         "Requires established track record"),
    ]
    for r, label, rates, note in wr_data:
        sc(ws, r, 2, label, font=LABEL_FONT)
        for i, v in enumerate(rates):
            sc(ws, r, 3 + i, v, font=INPUT_FONT, fmt=FMT_PCT, align=CENTER)
        sc(ws, r, 9, note, font=NOTE_FONT, align=LEFT_WRAP)

    # Renewal rate (row 24)
    sc(ws, 24, 2, "Recompete/Renewal Rate", font=LABEL_FONT)
    sc(ws, 24, 3, 0.70, font=INPUT_FONT, fmt=FMT_PCT, align=CENTER)
    sc(ws, 24, 9, "70% of contracts renew to incumbent", font=NOTE_FONT, align=LEFT_WRAP)

    # ── COST & MARGIN ASSUMPTIONS ──────────────────────────────
    banner(ws, 26, 2, 9, "COST & MARGIN ASSUMPTIONS")

    sc(ws, 27, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 27, 3, "Value", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(4, 9):
        sc(ws, 27, c, None, fill=LGRAY_FILL)
    sc(ws, 27, 9, "Notes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    sc(ws, 28, 2, "Blended Gross Margin", font=LABEL_FONT)
    sc(ws, 28, 3, 0.22, font=INPUT_FONT, fmt=FMT_PCT, align=CENTER)
    for c in range(4, 9):
        sc(ws, 28, c, None)
    ws.merge_cells(start_row=28, start_column=3, end_row=28, end_column=8)
    sc(ws, 28, 9, "22% weighted avg across tiers", font=NOTE_FONT, align=LEFT_WRAP)

    sc(ws, 29, 2, "Bid Prep Cost (Simplified+ per bid)", font=LABEL_FONT)
    sc(ws, 29, 3, 750, font=INPUT_FONT, fmt=FMT_CUR, align=CENTER)
    for c in range(4, 9):
        sc(ws, 29, c, None)
    ws.merge_cells(start_row=29, start_column=3, end_row=29, end_column=8)
    sc(ws, 29, 9, "Staff time + materials per RFQ/RFP", font=NOTE_FONT, align=LEFT_WRAP)

    sc(ws, 30, 2, "Fulfillment Overhead (% of rev)", font=LABEL_FONT)
    sc(ws, 30, 3, 0.03, font=INPUT_FONT, fmt=FMT_PCT, align=CENTER)
    for c in range(4, 9):
        sc(ws, 30, c, None)
    ws.merge_cells(start_row=30, start_column=3, end_row=30, end_column=8)
    sc(ws, 30, 9, "Compliance, logistics, admin", font=NOTE_FONT, align=LEFT_WRAP)

    # ── RENEWAL AVG CONTRACT VALUE ─────────────────────────────
    banner(ws, 32, 2, 9, "RENEWAL AVG CONTRACT VALUE")

    sc(ws, 33, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL)
    for i, h in enumerate(["Yr 1", "Yr 2", "Yr 3", "Yr 4", "Yr 5"]):
        sc(ws, 33, 3 + i, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 33, 8, None, fill=LGRAY_FILL)
    sc(ws, 33, 9, "Logic", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    sc(ws, 34, 2, "Avg Renewal Contract Value", font=LABEL_FONT)
    renewal_vals = [8000, 15000, 25000, 35000, 42000]
    for i, v in enumerate(renewal_vals):
        sc(ws, 34, 3 + i, v, font=INPUT_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 34, 9, "Early = micro renewals. Later = simplified+ renewals",
       font=NOTE_FONT, align=LEFT_WRAP)

    # ── ANNUAL PROGRAM INVESTMENT ──────────────────────────────
    banner(ws, 36, 2, 9, "ANNUAL PROGRAM INVESTMENT")

    sc(ws, 37, 2, "Item", font=LABEL_BOLD, fill=LGRAY_FILL)
    for i, h in enumerate(["Yr 1", "Yr 2", "Yr 3", "Yr 4", "Yr 5"]):
        sc(ws, 37, 3 + i, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 37, 8, "Total", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 37, 9, "Includes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    sc(ws, 38, 2, "Program Investment", font=LABEL_FONT)
    inv_vals = [11000, 6000, 5000, 4000, 4000]
    for i, v in enumerate(inv_vals):
        sc(ws, 38, 3 + i, v, font=INPUT_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 38, 8, "=SUM(C38:G38)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 38, 9, "Assoc dues, conferences, travel, software, insurance adj, legal review",
       font=NOTE_FONT, align=LEFT_WRAP)

    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Two Routes
# ---------------------------------------------------------------------------
def build_two_routes(wb):
    ws = wb.create_sheet("Two Routes")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 32

    title_rows(ws, 7)

    # ── COST COMPARISON ────────────────────────────────────────
    banner(ws, 5, 2, 7, "COST COMPARISON")

    # Header
    for c, h in [(2, "Item"), (3, "FREE ROUTE"), (4, ""), (5, "PAID ROUTE"), (6, ""), (7, "WHY")]:
        sc(ws, 6, c, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    cost_items = [
        ("SAM.gov Registration", "$0", "$0", "Required - both routes"),
        ("FL MyFloridaMarketplace", "$0", "$0", "Required for SLED"),
        ("Capability Statement", "DIY ($0)", "$500-$2,000", "Professional version wins more"),
        ("Food Safety Certification", "Existing ($0)", "$0-$23,500", "$0 if already certified"),
        ("Attorney Review (2-3 contracts)", "Skip", "$2,500-$5,000", "Gov contracts differ from commercial"),
        ("Insurance Adjustment", "Current ($0)", "$1,500-$3,000", "Some contracts need higher limits"),
        ("Industry Associations (NRA/IFDA)", "Skip", "$1,000-$2,500", "Bid alerts + networking access"),
        ("Trade Shows (2 events)", "Skip", "$3,000-$6,000", "Face time with contracting officers"),
        ("Relationship Building", "Cold calls ($0)", "$2,000-$5,000", "CO meetings, base/facility visits"),
        ("Bid Tracking Software", "Manual / free", "$500-$1,500", "GovWin IQ or similar"),
    ]

    for i, (item, free, paid, why) in enumerate(cost_items):
        r = 7 + i
        f = zebra(r)
        sc(ws, r, 2, item, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, free, font=GREEN_FONT, fill=f, align=CENTER)
        sc(ws, r, 4, None, fill=f)
        sc(ws, r, 5, paid, font=BLUE_FONT, fill=f, align=CENTER)
        sc(ws, r, 6, None, fill=f)
        sc(ws, r, 7, why, font=NOTE_FONT, fill=f, align=LEFT_WRAP)

    # Total row
    sc(ws, 17, 2, "TOTAL YEAR 1", font=TOTAL_FONT, fill=YELLOW_FILL)
    sc(ws, 17, 3, "$0 - $2,000", font=GREEN_BOLD, fill=YELLOW_FILL, align=CENTER)
    sc(ws, 17, 4, None, fill=YELLOW_FILL)
    sc(ws, 17, 5, "$11,000 - $47,000", font=BLUE_BOLD, fill=YELLOW_FILL, align=CENTER)
    sc(ws, 17, 6, None, fill=YELLOW_FILL)
    sc(ws, 17, 7, "No consulting fee - tools/hard costs only",
       font=TOTAL_FONT, fill=YELLOW_FILL, align=LEFT_WRAP)

    # ── EXPECTED OUTCOMES ──────────────────────────────────────
    banner(ws, 19, 2, 7, "EXPECTED OUTCOMES")

    for c, h in [(2, "Metric"), (3, "Free"), (4, ""), (5, "Paid"), (6, ""), (7, "Delta")]:
        sc(ws, 20, c, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    outcomes = [
        ("Yr 1 Win Rate (Micro)", "8-10%", "12-18%", "+50-80%"),
        ("Yr 1 Contracts Won", "4-6", "10-15", "2-3x more"),
        ("Yr 1 Revenue", "$25K-$45K", "$80K-$150K", "3-5x"),
        ("Time to Simplified Eligibility", "18-24 months", "6-9 months", "1+ year faster"),
        ("5-Yr Cumulative Revenue", "$400K-$800K", "$3M-$6M+", "8-10x"),
        ("5-Yr Cumulative Owner Earnings", "$50K-$120K", "$500K-$900K+", "10x+"),
    ]

    for i, (metric, free, paid, delta) in enumerate(outcomes):
        r = 21 + i
        f = zebra(r)
        sc(ws, r, 2, metric, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, free, font=GREEN_FONT, fill=f, align=CENTER)
        sc(ws, r, 4, None, fill=f)
        sc(ws, r, 5, paid, font=BLUE_BOLD, fill=f, align=CENTER)
        sc(ws, r, 6, None, fill=f)
        sc(ws, r, 7, delta, font=TOTAL_FONT, fill=f, align=CENTER)

    # Recommendation
    rec_text = (
        "RECOMMENDATION: Paid route. Year 1 is a small net investment that pays for "
        "itself by mid-Year 2. The compounding renewal base means every Year 1 dollar "
        "multiplies through Years 2-5."
    )
    sc(ws, 28, 2, rec_text, font=TOTAL_FONT, align=LEFT_WRAP)
    for c in range(3, 8):
        sc(ws, 28, c, None)
    ws.merge_cells(start_row=28, start_column=2, end_row=28, end_column=7)
    ws.row_dimensions[28].height = 40

    return ws


# ---------------------------------------------------------------------------
# Sheet 3: 5-Year Model
# ---------------------------------------------------------------------------
def build_five_year_model(wb):
    ws = wb.create_sheet("5-Year Model")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 36
    for c_letter in "CDEFGH":
        ws.column_dimensions[c_letter].width = 14
    ws.column_dimensions["I"].width = 28

    title_rows(ws, 9)

    yr_cols = {1: "C", 2: "D", 3: "E", 4: "F", 5: "G"}
    yr_col_nums = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7}

    def yr_header(row):
        sc(ws, row, 2, None, font=LABEL_BOLD, fill=LGRAY_FILL)
        for yr in range(1, 6):
            sc(ws, row, yr_col_nums[yr], f"Yr {yr}",
               font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
        sc(ws, row, 8, "Total", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
        sc(ws, row, 9, "Strategy", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    # ═══════════════════════════════════════════════════════════
    # NEW BIDS SUBMITTED (rows 5-13)
    # ═══════════════════════════════════════════════════════════
    banner(ws, 5, 2, 9, "NEW BIDS SUBMITTED")

    sc(ws, 6, 2, "Blue values are editable assumptions", font=NOTE_FONT)
    for c in range(3, 10):
        sc(ws, 6, c, None)
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=9)

    yr_header(7)

    # Bid counts (editable blue values)
    bids = [
        (8,  "Micro (<$15K) CREDIBILITY PHASE",
         [55, 25, 10, 0, 0], "Win enough to unlock, then stop actively bidding"),
        (9,  "Simplified ($15K-$250K)",
         [8, 28, 42, 52, 58], "Starts Q3 Y1 w/ past perf. Primary engine Y2+"),
        (10, "Set-Aside (SDB/HUBZone)",
         [6, 14, 22, 28, 30], "Day 1 if certified. Premium margins"),
        (11, "SLED (State/Local/Ed)",
         [15, 22, 30, 38, 42], "FL schools/counties - parallel from Day 1"),
        (12, "Sealed Bid ($250K+)",
         [0, 3, 8, 12, 16], "Y2+ only with track record"),
    ]

    for r, label, vals, strategy in bids:
        lbl_font = ORANGE_BOLD if r == 8 else LABEL_FONT
        sc(ws, r, 2, label, font=lbl_font)
        for i, v in enumerate(vals):
            sc(ws, r, 3 + i, v, font=INPUT_FONT, fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 8, f"=SUM(C{r}:G{r})", font=TOTAL_FONT, fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 9, strategy, font=NOTE_FONT, align=LEFT_WRAP)

    # Total bids row
    sc(ws, 13, 2, "TOTAL BIDS", font=TOTAL_FONT, fill=YELLOW_FILL)
    for yr in range(1, 6):
        c = yr_col_nums[yr]
        cl = yr_cols[yr]
        sc(ws, 13, c, f"=SUM({cl}8:{cl}12)", font=TOTAL_FONT, fill=YELLOW_FILL,
           fmt=FMT_NUM, align=CENTER)
    sc(ws, 13, 8, "=SUM(H8:H12)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_NUM, align=CENTER)
    sc(ws, 13, 9, None, fill=YELLOW_FILL)

    # ═══════════════════════════════════════════════════════════
    # NEW CONTRACT WINS (rows 15-22)
    # ═══════════════════════════════════════════════════════════
    banner(ws, 15, 2, 9, "NEW CONTRACT WINS")

    yr_header(16)

    # Win formulas: =ROUND(bids * Inputs!win_rate, 0)
    # Inputs win rate rows: 19=Micro, 20=Simplified, 21=SetAside, 22=SLED, 23=Sealed
    win_tiers = [
        (17, "Micro (<$15K)", 8, 19),
        (18, "Simplified ($15K-$250K)", 9, 20),
        (19, "Set-Aside (SDB/HUBZone)", 10, 21),
        (20, "SLED (State/Local/Ed)", 11, 22),
        (21, "Sealed Bid ($250K+)", 12, 23),
    ]
    for r, label, bid_row, wr_row in win_tiers:
        sc(ws, r, 2, label, font=LABEL_FONT)
        for yr in range(1, 6):
            cl = yr_cols[yr]
            sc(ws, r, yr_col_nums[yr],
               f"=ROUND({cl}{bid_row}*Inputs!{cl}{wr_row},0)",
               font=FORMULA_FONT, fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 8, f"=SUM(C{r}:G{r})", font=TOTAL_FONT, fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 9, None)

    # Total wins
    sc(ws, 22, 2, "TOTAL NEW WINS", font=TOTAL_FONT, fill=YELLOW_FILL)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 22, yr_col_nums[yr], f"=SUM({cl}17:{cl}21)",
           font=TOTAL_FONT, fill=YELLOW_FILL, fmt=FMT_NUM, align=CENTER)
    sc(ws, 22, 8, "=SUM(H17:H21)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_NUM, align=CENTER)
    sc(ws, 22, 9, None, fill=YELLOW_FILL)

    # ═══════════════════════════════════════════════════════════
    # ACTIVE CONTRACT BASE (rows 24-29)
    # ═══════════════════════════════════════════════════════════
    banner(ws, 24, 2, 9, "ACTIVE CONTRACT BASE")

    sc(ws, 25, 2, "70% renewal rate compounds the portfolio each year", font=NOTE_FONT)
    for c in range(3, 10):
        sc(ws, 25, c, None)
    ws.merge_cells(start_row=25, start_column=2, end_row=25, end_column=9)

    yr_header(26)

    # Renewed from prior year (row 27)
    sc(ws, 27, 2, "Renewed from Prior Year (70%)", font=LABEL_FONT)
    sc(ws, 27, 3, 0, font=FORMULA_FONT, fmt=FMT_NUM, align=CENTER)  # Yr1 = 0
    for yr in range(2, 6):
        prev_cl = yr_cols[yr - 1]
        sc(ws, 27, yr_col_nums[yr],
           f"=ROUND({prev_cl}29*Inputs!$C$24,0)",
           font=FORMULA_FONT, fmt=FMT_NUM, align=CENTER)
    sc(ws, 27, 8, "=SUM(C27:G27)", font=TOTAL_FONT, fmt=FMT_NUM, align=CENTER)
    sc(ws, 27, 9, None)

    # New wins this year (row 28)
    sc(ws, 28, 2, "+ New Wins This Year", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 28, yr_col_nums[yr], f"={cl}22",
           font=FORMULA_FONT, fmt=FMT_NUM, align=CENTER)
    sc(ws, 28, 8, "=SUM(C28:G28)", font=TOTAL_FONT, fmt=FMT_NUM, align=CENTER)
    sc(ws, 28, 9, None)

    # Total active (row 29)
    sc(ws, 29, 2, "TOTAL ACTIVE CONTRACTS", font=TOTAL_FONT, fill=YELLOW_FILL)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 29, yr_col_nums[yr], f"={cl}27+{cl}28",
           font=TOTAL_FONT, fill=YELLOW_FILL, fmt=FMT_NUM, align=CENTER)
    sc(ws, 29, 8, "=SUM(C29:G29)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_NUM, align=CENTER)
    sc(ws, 29, 9, None, fill=YELLOW_FILL)

    # ═══════════════════════════════════════════════════════════
    # REVENUE (rows 31-39)
    # ═══════════════════════════════════════════════════════════
    banner(ws, 31, 2, 9, "REVENUE")

    sc(ws, 32, 2, "Source", font=LABEL_BOLD, fill=LGRAY_FILL)
    for yr in range(1, 6):
        sc(ws, 32, yr_col_nums[yr], f"Yr {yr}",
           font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 32, 8, "5-Yr Total", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 32, 9, "Avg Contract", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    # Revenue lines: wins_row × avg_contract_value (in col I)
    rev_tiers = [
        (33, "Micro Wins (new)", 17, 5200),
        (34, "Simplified Wins (new)", 18, 45000),
        (35, "Set-Aside Wins (new)", 19, 40000),
        (36, "SLED Wins (new)", 20, 18000),
        (37, "Sealed Bid Wins (new)", 21, 175000),
    ]
    for r, label, win_row, avg_val in rev_tiers:
        sc(ws, r, 2, label, font=LABEL_FONT)
        sc(ws, r, 9, avg_val, font=NOTE_FONT, fmt=FMT_CUR, align=CENTER)
        for yr in range(1, 6):
            cl = yr_cols[yr]
            sc(ws, r, yr_col_nums[yr], f"={cl}{win_row}*$I${r}",
               font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 8, f"=SUM(C{r}:G{r})", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # Renewal revenue (row 38)
    sc(ws, 38, 2, "Renewal Revenue (recurring)", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 38, yr_col_nums[yr], f"={cl}27*Inputs!{cl}34",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 38, 8, "=SUM(C38:G38)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 38, 9, "Avg grows as portfolio shifts up", font=NOTE_FONT, align=LEFT_WRAP)

    # Total revenue (row 39)
    sc(ws, 39, 2, "TOTAL REVENUE", font=TOTAL_FONT, fill=YELLOW_FILL)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 39, yr_col_nums[yr], f"=SUM({cl}33:{cl}38)",
           font=TOTAL_FONT, fill=YELLOW_FILL, fmt=FMT_CUR, align=CENTER)
    sc(ws, 39, 8, "=SUM(C39:G39)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_CUR, align=CENTER)
    sc(ws, 39, 9, None, fill=YELLOW_FILL)

    # ═══════════════════════════════════════════════════════════
    # OWNER EARNINGS (rows 41-52)
    # ═══════════════════════════════════════════════════════════
    banner(ws, 41, 2, 9, "OWNER EARNINGS")

    sc(ws, 42, 2, "What Newport keeps after all costs", font=NOTE_FONT)
    for c in range(3, 10):
        sc(ws, 42, c, None)
    ws.merge_cells(start_row=42, start_column=2, end_row=42, end_column=9)

    sc(ws, 43, 2, "Line Item", font=LABEL_BOLD, fill=LGRAY_FILL)
    for yr in range(1, 6):
        sc(ws, 43, yr_col_nums[yr], f"Yr {yr}",
           font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 43, 8, "5-Yr Total", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 43, 9, None, fill=LGRAY_FILL)

    # Revenue (row 44)
    sc(ws, 44, 2, "Revenue", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 44, yr_col_nums[yr], f"={cl}39",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 44, 8, "=SUM(C44:G44)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # COGS (row 45)
    sc(ws, 45, 2, "(-) COGS", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 45, yr_col_nums[yr], f"={cl}44*(1-Inputs!$C$28)",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 45, 8, "=SUM(C45:G45)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # Gross Profit (row 46)
    sc(ws, 46, 2, "Gross Profit (22%)", font=LABEL_BOLD)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 46, yr_col_nums[yr], f"={cl}44-{cl}45",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 46, 8, "=SUM(C46:G46)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # Bid Prep (row 47) - non-micro bids only (rows 9-12)
    sc(ws, 47, 2, "(-) Bid Prep ($750/bid)", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 47, yr_col_nums[yr],
           f"=({cl}9+{cl}10+{cl}11+{cl}12)*Inputs!$C$29",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 47, 8, "=SUM(C47:G47)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # Fulfillment (row 48)
    sc(ws, 48, 2, "(-) Fulfillment/Compliance (3%)", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 48, yr_col_nums[yr], f"={cl}44*Inputs!$C$30",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 48, 8, "=SUM(C48:G48)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # Program Costs (row 49)
    sc(ws, 49, 2, "(-) Program Costs", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 49, yr_col_nums[yr], f"=Inputs!{cl}38",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 49, 8, "=SUM(C49:G49)", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # Owner Earnings (row 50)
    sc(ws, 50, 2, "OWNER EARNINGS", font=TOTAL_FONT, fill=YELLOW_FILL)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 50, yr_col_nums[yr], f"={cl}46-{cl}47-{cl}48-{cl}49",
           font=TOTAL_FONT, fill=YELLOW_FILL, fmt=FMT_CUR, align=CENTER)
    sc(ws, 50, 8, "=SUM(C50:G50)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_CUR, align=CENTER)
    sc(ws, 50, 9, None, fill=YELLOW_FILL)

    # Cumulative OE (row 51)
    sc(ws, 51, 2, "Cumulative Owner Earnings", font=LABEL_FONT)
    sc(ws, 51, 3, "=C50", font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    for yr in range(2, 6):
        cl = yr_cols[yr]
        prev_cl = yr_cols[yr - 1]
        sc(ws, 51, yr_col_nums[yr], f"={prev_cl}51+{cl}50",
           font=FORMULA_FONT, fmt=FMT_CUR, align=CENTER)
    sc(ws, 51, 8, "=G51", font=TOTAL_FONT, fmt=FMT_CUR, align=CENTER)

    # OE Margin (row 52)
    sc(ws, 52, 2, "OE Margin", font=LABEL_FONT)
    for yr in range(1, 6):
        cl = yr_cols[yr]
        sc(ws, 52, yr_col_nums[yr], f"=IF({cl}44>0,{cl}50/{cl}44,0)",
           font=FORMULA_FONT, fmt=FMT_PCT, align=CENTER)
    sc(ws, 52, 8, None)

    # ═══════════════════════════════════════════════════════════
    # CHARTS (rows 55-112)
    # ═══════════════════════════════════════════════════════════

    # Chart 1: Revenue by Source (stacked bar)
    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "stacked"
    chart1.title = "Revenue by Source - Renewals Compound Each Year"
    chart1.y_axis.title = "Revenue ($)"
    chart1.x_axis.title = None
    chart1.style = 10
    chart1.width = 20
    chart1.height = 12
    cats = Reference(ws, min_col=3, max_col=7, min_row=32)  # Yr 1-5 headers
    for rev_row in range(33, 39):  # 6 revenue sources
        data = Reference(ws, min_col=3, max_col=7, min_row=rev_row)
        chart1.add_data(data, from_rows=True, titles_from_data=False)
        chart1.series[-1].tx = SeriesLabel(v=ws.cell(row=rev_row, column=2).value or "")
    chart1.set_categories(cats)
    ws.add_chart(chart1, "B55")

    # Chart 2: Owner Earnings annual + cumulative (bar)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Owner Earnings: Annual (blue) + Cumulative (gold)"
    chart2.style = 10
    chart2.width = 20
    chart2.height = 12
    cats2 = Reference(ws, min_col=3, max_col=7, min_row=43)
    data_oe = Reference(ws, min_col=3, max_col=7, min_row=50)
    data_cum = Reference(ws, min_col=3, max_col=7, min_row=51)
    chart2.add_data(data_oe, from_rows=True, titles_from_data=False)
    chart2.series[-1].tx = SeriesLabel(v="Annual OE")
    chart2.add_data(data_cum, from_rows=True, titles_from_data=False)
    chart2.series[-1].tx = SeriesLabel(v="Cumulative OE")
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "B72")

    # Chart 3: Active contracts (stacked bar - renewals + new)
    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "stacked"
    chart3.title = "Renewals (gold) Stack on New Wins (blue) Every Year"
    chart3.style = 10
    chart3.width = 20
    chart3.height = 12
    cats3 = Reference(ws, min_col=3, max_col=7, min_row=26)
    data_ren = Reference(ws, min_col=3, max_col=7, min_row=27)
    data_new = Reference(ws, min_col=3, max_col=7, min_row=28)
    chart3.add_data(data_ren, from_rows=True, titles_from_data=False)
    chart3.series[-1].tx = SeriesLabel(v="Renewed")
    chart3.add_data(data_new, from_rows=True, titles_from_data=False)
    chart3.series[-1].tx = SeriesLabel(v="New Wins")
    chart3.set_categories(cats3)
    ws.add_chart(chart3, "B89")

    # ═══════════════════════════════════════════════════════════
    # PORTFOLIO SHIFT (rows 113-120)
    # ═══════════════════════════════════════════════════════════
    banner(ws, 113, 2, 9, "PORTFOLIO SHIFT")

    sc(ws, 114, 2, "Tier", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 114, 3, "Year 1 %", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 114, 4, "Year 5 %", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(5, 10):
        sc(ws, 114, c, None, fill=LGRAY_FILL)

    shift_tiers = [
        (115, "Micro", 33),
        (116, "Simplified", 34),
        (117, "Set-Aside", 35),
        (118, "SLED", 36),
        (119, "Sealed Bid", 37),
        (120, "Renewals", 38),
    ]
    for r, label, rev_row in shift_tiers:
        sc(ws, r, 2, label, font=LABEL_FONT)
        sc(ws, r, 3, f"=IF(C39>0,C{rev_row}/C39,0)",
           font=FORMULA_FONT, fmt=FMT_PCT, align=CENTER)
        sc(ws, r, 4, f"=IF(G39>0,G{rev_row}/G39,0)",
           font=FORMULA_FONT, fmt=FMT_PCT, align=CENTER)

    # Pie charts for portfolio shift
    pie1 = PieChart()
    pie1.title = "Year 1: Credibility Phase"
    pie1.style = 10
    pie1.width = 14
    pie1.height = 10
    pie1_data = Reference(ws, min_col=3, min_row=115, max_row=120)
    pie1_cats = Reference(ws, min_col=2, min_row=115, max_row=120)
    pie1.add_data(pie1_data, titles_from_data=False)
    pie1.set_categories(pie1_cats)
    ws.add_chart(pie1, "F114")

    pie2 = PieChart()
    pie2.title = "Year 5: Diversified, Micro Gone"
    pie2.style = 10
    pie2.width = 14
    pie2.height = 10
    pie2_data = Reference(ws, min_col=4, min_row=115, max_row=120)
    pie2_cats = Reference(ws, min_col=2, min_row=115, max_row=120)
    pie2.add_data(pie2_data, titles_from_data=False)
    pie2.set_categories(pie2_cats)
    ws.add_chart(pie2, "F126")

    return ws


# ---------------------------------------------------------------------------
# Sheet 4: Market Analysis
# ---------------------------------------------------------------------------
def build_market_analysis(wb):
    ws = wb.create_sheet("Market Analysis")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 28

    title_rows(ws, 8)

    # ── ADDRESSABLE MARKET WATERFALL ───────────────────────────
    banner(ws, 5, 2, 8, "ADDRESSABLE MARKET WATERFALL")

    headers = ["Level", "Scope", "Annual $", "# Awards", "Avg Size", "Newport Fit"]
    for i, h in enumerate(headers):
        sc(ws, 6, 2 + i, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 8, None, fill=LGRAY_FILL)

    tam_data = [
        ("Federal Food (National)", "All NAICS 311/424", 7200000000, "~500K",
         "$14.4K", "NOT REALISTIC", RED_BOLD),
        ("Federal (<$350K)", "Realistic bid range", 2100000000, "~300K",
         "$7K", "Too broad", RED_FONT),
        ("Florida Federal (<$350K)", "Home state", 85000000, "39,685",
         "$2.1K", "ADDRESSABLE", GREEN_BOLD),
        ("Southeast (<$350K)", "FL+GA+SC+AL+MS", 179000000, "~80K",
         "$2.2K", "IF multi-state", ORANGE_FONT),
        ("FL + SLED Combined", "Fed + State/Local", 135000000, "~55K",
         "$2.5K", "TOTAL OPP", GREEN_BOLD),
    ]

    for i, (level, scope, annual, awards, avg, fit, fit_font) in enumerate(tam_data):
        r = 7 + i
        f = zebra(r)
        sc(ws, r, 2, level, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, scope, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 4, annual, font=FORMULA_FONT, fill=f, fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 5, awards, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 6, avg, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 7, fit, font=fit_font, fill=f, align=CENTER)
        sc(ws, r, 8, None, fill=f)

    # Chart data for waterfall bar chart (rows 13-18)
    sc(ws, 13, 2, "Chart Data", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 13, 3, "$M Annual", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    chart_bars = [
        (14, "National (<$350K)", 2100),
        (15, "Florida", 85),
        (16, "Southeast", 179),
        (17, "FL+SLED", 135),
        (18, "Confectionery Gap", 55),
    ]
    for r, label, val in chart_bars:
        sc(ws, r, 2, label, font=LABEL_FONT)
        sc(ws, r, 3, val, font=FORMULA_FONT, fmt=FMT_NUM, align=CENTER)

    # Market waterfall chart
    chart_mkt = BarChart()
    chart_mkt.type = "col"
    chart_mkt.title = "Market Waterfall ($M Annual)"
    chart_mkt.style = 10
    chart_mkt.width = 20
    chart_mkt.height = 12
    mkt_data = Reference(ws, min_col=3, min_row=14, max_row=18)
    mkt_cats = Reference(ws, min_col=2, min_row=14, max_row=18)
    chart_mkt.add_data(mkt_data, titles_from_data=False)
    chart_mkt.set_categories(mkt_cats)
    ws.add_chart(chart_mkt, "E6")

    # ── REAL CONTRACT EXAMPLES ─────────────────────────────────
    banner(ws, 34, 2, 8, "REAL CONTRACT EXAMPLES")

    ex_headers = ["Tier", "Agency", "Value", "Location", "Product", "Frequency"]
    for i, h in enumerate(ex_headers):
        sc(ws, 35, 2 + i, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 35, 8, None, fill=LGRAY_FILL)

    examples = [
        ("Micro", "USDA Forest Service", "$3,800", "Ocala NF, FL", "Trail mix/snacks", "Quarterly"),
        ("Micro", "VA Medical Center", "$4,200", "Bay Pines, FL", "Candy/confectionery", "Monthly"),
        ("Micro", "DeCA Commissary", "$8,500", "MacDill AFB, FL", "Snack assortment", "Monthly"),
        ("Micro", "DOD Elementary", "$5,100", "Patrick SFB, FL", "Snack packs", "Weekly"),
        ("Set-Aside", "BOP Coleman", "$28,000", "Sumterville, FL", "Commissary snacks", "Monthly"),
        ("Set-Aside", "BOP Miami FDC", "$35,000", "Miami, FL", "Dry goods bundle", "Monthly"),
        ("Simplified", "School District", "$65,000", "Broward County", "Snack program", "Annual"),
        ("Simplified", "State DOC", "$85,000", "FL Corrections", "Commissary supply", "Annual"),
        ("SLED", "County Gov", "$42,000", "Palm Beach, FL", "Event concessions", "Seasonal"),
        ("SLED", "State University", "$55,000", "FSU, Tallahassee", "Vending supply", "Annual"),
        ("Sealed Bid", "DeCA Regional", "$220,000", "SE Region", "Full confectionery line", "Annual"),
    ]

    for i, (tier, agency, val, loc, prod, freq) in enumerate(examples):
        r = 36 + i
        f = zebra(r)
        ws.row_dimensions[r].height = 24
        sc(ws, r, 2, tier, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 3, agency, font=LABEL_FONT, fill=f)
        sc(ws, r, 4, val, font=INPUT_FONT, fill=f, align=CENTER)
        sc(ws, r, 5, loc, font=LABEL_FONT, fill=f)
        sc(ws, r, 6, prod, font=LABEL_FONT, fill=f)
        sc(ws, r, 7, freq, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 8, None, fill=f)

    # ── CATEGORY ADVANTAGE ─────────────────────────────────────
    banner(ws, 48, 2, 8, "CATEGORY ADVANTAGE")

    cat_headers = ["PSC Category", "National TAM", "FL Competition", "Gap", "Priority"]
    for i, h in enumerate(cat_headers):
        sc(ws, 49, 2 + i, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(7, 9):
        sc(ws, 49, c, None, fill=LGRAY_FILL)

    categories = [
        ("8925 Confectionery/Nuts", 55000000, 412000, "WIDE OPEN", GREEN_BOLD),
        ("8940 Snack Foods", 38000000, 2100000, "Strong", BLUE_BOLD),
        ("8950 Bakery/Cereal", 31000000, 950000, "Good", BLUE_FONT),
        ("8945 Beverages (non-alc)", 22000000, 1800000, "Moderate", BLUE_FONT),
    ]

    for i, (cat, tam, comp, priority, pri_font) in enumerate(categories):
        r = 50 + i
        f = zebra(r)
        sc(ws, r, 2, cat, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, tam, font=FORMULA_FONT, fill=f, fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 4, comp, font=FORMULA_FONT, fill=f, fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 5, f"=C{r}-D{r}", font=FORMULA_FONT, fill=f, fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 6, priority, font=pri_font, fill=f, align=CENTER)
        for c in range(7, 9):
            sc(ws, r, c, None, fill=f)

    # Category chart (grouped bar)
    chart_cat = BarChart()
    chart_cat.type = "col"
    chart_cat.grouping = "clustered"
    chart_cat.title = "National TAM vs FL Competition by Category"
    chart_cat.style = 10
    chart_cat.width = 20
    chart_cat.height = 12
    cat_cats = Reference(ws, min_col=2, min_row=50, max_row=53)
    cat_tam = Reference(ws, min_col=3, min_row=50, max_row=53)
    cat_comp = Reference(ws, min_col=4, min_row=50, max_row=53)
    chart_cat.add_data(cat_tam, titles_from_data=False)
    chart_cat.series[-1].tx = SeriesLabel(v="National TAM")
    chart_cat.add_data(cat_comp, titles_from_data=False)
    chart_cat.series[-1].tx = SeriesLabel(v="FL Competition")
    chart_cat.set_categories(cat_cats)
    ws.add_chart(chart_cat, "B55")

    # ── COMPETITIVE LANDSCAPE ──────────────────────────────────
    banner(ws, 70, 2, 8, "COMPETITIVE LANDSCAPE")

    comp_headers = ["Competitor", "Gov Rev", "Categories", "Strength", "Weakness", "Threat"]
    for i, h in enumerate(comp_headers):
        sc(ws, 71, 2 + i, h, font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(8, 9):
        sc(ws, 71, c, None, fill=LGRAY_FILL)

    competitors = [
        ("Oakes Farms", "$26M", "Produce/meats", "Established", "No confectionery", "LOW", GREEN_FONT),
        ("US Foods", "$24M", "Full-line", "Scale", "Ignores micro", "LOW", GREEN_FONT),
        ("Rainmaker Inc.", "$5M", "DOJ/BOP", "Incumbent", "Narrow catalog", "TARGET", RED_BOLD),
        ("Sysco Gov", "$18M", "Institutional", "National", "Not micro-focused", "LOW", GREEN_FONT),
        ("Local Distributors", "$8M", "Mixed", "Relationships", "No scale", "MEDIUM", ORANGE_FONT),
    ]

    for i, (name, rev, cats, strength, weakness, threat, t_font) in enumerate(competitors):
        r = 72 + i
        f = zebra(r)
        sc(ws, r, 2, name, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, rev, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 4, cats, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 5, strength, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 6, weakness, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 7, threat, font=t_font, fill=f, align=CENTER)
        sc(ws, r, 8, None, fill=f)

    # Sources
    sc(ws, 78, 2, "Sources: USASpending.gov API, FPDS-NG, SAM.gov, FL "
       "MyFloridaMarketplace. Contract examples are representative composites "
       "based on actual award data. Feb 2026.", font=NOTE_FONT)
    for c in range(3, 9):
        sc(ws, 78, c, None)
    ws.merge_cells(start_row=78, start_column=2, end_row=78, end_column=8)

    return ws


# ---------------------------------------------------------------------------
# Sheet 5: Key Questions
# ---------------------------------------------------------------------------
def build_key_questions(wb):
    ws = wb.create_sheet("Key Questions")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 32

    title_rows(ws, 5)

    banner(ws, 5, 2, 5, "CRITICAL DECISIONS THAT CHANGE THE MODEL")

    sc(ws, 6, 2, "Priority", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 3, "Question", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 4, "If Yes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 5, "If No", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    questions = [
        ("1 HIGHEST", RED_BOLD,
         "Can Newport profitably fulfill orders under $5,200?",
         "Micro channel opens (83% of awards). Credibility ramp starts.",
         "Must skip to Simplified - harder, slower entry path"),
        ("2 HIGHEST", RED_BOLD,
         "Eligible for SDB/HUBZone/8(a) set-aside?",
         "Win rates 28%+, margins 25%. Available from Day 1.",
         "Open competition only - lower win rates, slower ramp"),
        ("3 HIGH", ORANGE_BOLD,
         "Delivery radius from FL warehouses?",
         "FL only = $85M. SE region = $179M. Determines total addressable.",
         "We right-size model to actual reach"),
        ("4 HIGH", ORANGE_BOLD,
         "Existing food safety certs (SQF/GFSI)?",
         "Year 1 investment drops by $23,500",
         "Biggest single cost variable"),
        ("5 HIGH", ORANGE_BOLD,
         "Willing to invest $11K-$47K (Paid Route)?",
         "Full competitive capability from Day 1. ROI positive by mid-Y2.",
         "Free route works but 12-18 months slower to traction"),
        ("6 MEDIUM", BLUE_BOLD,
         "How many confectionery/snack/nut SKUs?",
         "Broader catalog = eligible for more bids across categories",
         "We focus on strongest categories only"),
        ("7 MEDIUM", BLUE_BOLD,
         "Can add delivery routes to bases/prisons?",
         "DOJ/BOP channel opens ($5M+ FL). Rainmaker Inc. is beatable.",
         "Skip DOJ, focus DeCA commissaries instead"),
        ("8 MEDIUM", BLUE_BOLD,
         "Current commercial gross margin?",
         "Calibrates if gov margins (18-25%) are accretive or dilutive",
         "May not be worth pursuing if commercial margins much higher"),
        ("9 INFO", NOTE_FONT,
         "Any prior gov contract experience?",
         "Past performance references accelerate everything",
         "Start from zero - typical for new entrants"),
        ("10 INFO", NOTE_FONT,
         "Goal: supplemental revenue or primary channel?",
         "Aggressive staffing and investment trajectory",
         "Conservative side channel approach"),
    ]

    for i, (priority, pri_font, question, if_yes, if_no) in enumerate(questions):
        r = 7 + i
        f = zebra(r)
        ws.row_dimensions[r].height = 40
        sc(ws, r, 2, priority, font=pri_font, fill=f, align=CENTER)
        sc(ws, r, 3, question, font=LABEL_FONT, fill=f, align=LEFT_WRAP)
        sc(ws, r, 4, if_yes, font=GREEN_FONT, fill=f, align=LEFT_WRAP)
        sc(ws, r, 5, if_no, font=RED_FONT, fill=f, align=LEFT_WRAP)

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Build Newport GovCon Financial Model (v7)")
    parser.add_argument("--output", help="Output path (default: same directory)")
    args = parser.parse_args()

    output_dir = Path(__file__).resolve().parent
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = output_dir / "newport-govcon-financial-model.xlsx"

    print("=" * 60)
    print("Newport GovCon Financial Model Builder (v7)")
    print("=" * 60)

    wb = Workbook()

    print("[1/5] Building Inputs sheet...")
    build_inputs(wb)

    print("[2/5] Building Two Routes sheet...")
    build_two_routes(wb)

    print("[3/5] Building 5-Year Model sheet (with 5 charts)...")
    build_five_year_model(wb)

    print("[4/5] Building Market Analysis sheet (with 2 charts)...")
    build_market_analysis(wb)

    print("[5/5] Building Key Questions sheet...")
    build_key_questions(wb)

    wb.save(str(output_path))
    print()
    print(f"Saved: {output_path}")
    print(f"File size: {output_path.stat().st_size:,} bytes")
    print()

    print("Verification:")
    print(f"  Sheets: {wb.sheetnames}")
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"  {ws_name}: {ws.max_row} rows x {ws.max_column} cols")

    print()
    print("Design: Blue text = editable inputs | Formulas = calculated")
    print("Yellow rows = totals | Navy = titles | Blue banners = sections")
    print("7 charts total (5 in 5-Year Model, 2 in Market Analysis)")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
