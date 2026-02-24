#!/usr/bin/env python3
"""
Build Newport Commercial SDR Financial Model.

Generates a 5-sheet Excel workbook for Newport's commercial outbound
prospecting channel across 5 ICP segments with 3 investment scenarios:

  1. Inputs         - Outreach parameters, tool costs, capacity
  2. ICP Segments   - Per-segment economics for 5 segments
  3. Funnel Model   - 12-month projection x 3 scenarios with 3 charts
  4. Market Analysis - TAM by segment, candy market, tool matrix
  5. Key Questions   - 10 critical decisions

ALL calculations use Excel formulas, NOT Python math.

Usage:
    python commercial/deliverables/financials/build_commercial_model.py
    python commercial/deliverables/financials/build_commercial_model.py --output data/final/model.xlsx
"""

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Design system (matches GovCon model exactly)
# ---------------------------------------------------------------------------
# Colors
DARK_NAVY = "1B2A4A"
MEDIUM_BLUE = "2E5090"
PURE_BLUE = "0000FF"
GREEN = "548235"
ORANGE = "ED7D31"
RED = "C00000"
GRAY = "808080"
DARK_GRAY = "404040"
WHITE = "FFFFFF"
LIGHT_GRAY_HEX = "F2F2F2"
PALE_YELLOW = "FFF2CC"

# Fonts
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=11, color=WHITE)
BANNER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=11, color=DARK_GRAY)
LABEL_BOLD = Font(name="Calibri", bold=True, size=11, color=DARK_GRAY)
INPUT_FONT = Font(name="Calibri", size=11, color=PURE_BLUE)
INPUT_BOLD = Font(name="Calibri", bold=True, size=11, color=PURE_BLUE)
NOTE_FONT = Font(name="Calibri", size=10, color=GRAY, italic=True)
GREEN_FONT = Font(name="Calibri", size=11, color=GREEN)
GREEN_BOLD = Font(name="Calibri", bold=True, size=11, color=GREEN)
ORANGE_FONT = Font(name="Calibri", size=11, color=ORANGE)
ORANGE_BOLD = Font(name="Calibri", bold=True, size=11, color=ORANGE)
RED_FONT = Font(name="Calibri", size=11, color=RED)
RED_BOLD = Font(name="Calibri", bold=True, size=11, color=RED)
BLUE_FONT = Font(name="Calibri", size=11, color=MEDIUM_BLUE)
BLUE_BOLD = Font(name="Calibri", bold=True, size=11, color=MEDIUM_BLUE)
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color=DARK_GRAY)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
FORMULA_FONT = Font(name="Calibri", size=11, color="000000")

# Fills
NAVY_FILL = PatternFill("solid", fgColor=DARK_NAVY)
BLUE_FILL = PatternFill("solid", fgColor=MEDIUM_BLUE)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)
LGRAY_FILL = PatternFill("solid", fgColor=LIGHT_GRAY_HEX)
YELLOW_FILL = PatternFill("solid", fgColor=PALE_YELLOW)

# Borders
_THIN = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Number formats
FMT_CUR = '$#,##0;"($"#,##0);-'
FMT_PCT = "0.0%"
FMT_NUM = "#,##0"
FMT_CUR_M = '$#,##0,,"M"'

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sc(ws, r, c, val, font=None, fill=None, fmt=None, align=None):
    """Set cell value + style."""
    cell = ws.cell(row=r, column=c, value=val)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    cell.border = THIN_BORDER
    return cell


def banner(ws, row, col_start, col_end, text, fill=None):
    """Write a merged banner row."""
    f = fill or BLUE_FILL
    sc(ws, row, col_start, text, font=BANNER_FONT, fill=f, align=CENTER)
    for c in range(col_start + 1, col_end + 1):
        sc(ws, row, c, None, fill=f)
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end,
    )


def title_rows(ws, col_end):
    """Write standard title + subtitle in rows 2-3."""
    now = datetime.now()
    sc(ws, 2, 2, "Newport Wholesalers \u2014 Commercial SDR Model",
       font=TITLE_FONT, fill=NAVY_FILL, align=CENTER)
    for c in range(3, col_end + 1):
        sc(ws, 2, c, None, fill=NAVY_FILL)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=col_end)
    ws.row_dimensions[2].height = 20

    sc(ws, 3, 2, f"12-Month Projection | {now.strftime('%B %Y')}",
       font=SUBTITLE_FONT, fill=NAVY_FILL, align=CENTER)
    for c in range(3, col_end + 1):
        sc(ws, 3, c, None, fill=NAVY_FILL)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=col_end)


def zebra(row_num):
    """Return alternating fill for zebra striping."""
    return LGRAY_FILL if row_num % 2 == 0 else WHITE_FILL


# ---------------------------------------------------------------------------
# Data constants
# ---------------------------------------------------------------------------
SEGMENTS = [
    {
        "code": "A", "name": "Enterprise Grocery",
        "pipeline": "Direct-to-customer", "geo": "US/UK/EU",
        "companies": 2000, "apollo_cov": 0.65,
        "deal_size": 50000, "priority": 1,
        "notes": "Largest deal size",
    },
    {
        "code": "B", "name": "Manufacturers",
        "pipeline": "Partner", "geo": "US/UK/EU",
        "companies": 3000, "apollo_cov": 0.60,
        "deal_size": 25000, "priority": 2,
        "notes": "Volume play",
    },
    {
        "code": "C", "name": "Gov Buyers",
        "pipeline": "Direct-to-customer", "geo": "US only",
        "companies": 500, "apollo_cov": 0.40,
        "deal_size": 15000, "priority": 3,
        "notes": "Overlap with GovCon channel",
    },
    {
        "code": "D", "name": "Corrections",
        "pipeline": "Direct-to-customer", "geo": "US only",
        "companies": 200, "apollo_cov": 0.55,
        "deal_size": 30000, "priority": 3,
        "notes": "Niche, high barriers",
    },
    {
        "code": "E", "name": "Candy/Confectionery",
        "pipeline": "Partner", "geo": "US/LATAM",
        "companies": 1831, "apollo_cov": 0.60,
        "deal_size": 20000, "priority": 1,
        "notes": "Newport's specialty",
    },
]

SCENARIOS = ["Free", "Moderate", "Aggressive"]

# Tool costs by scenario (annual)
TOOL_COSTS = [
    # (tool_name, free, moderate, aggressive)
    ("Apollo.io", 0, 588, 1428),
    ("Instantly", 0, 444, 1164),
    ("Clay", 0, 0, 4680),
    ("Twilio SMS", 0, 0, 600),
    ("Retell AI Voice", 0, 0, 1200),
]

# Weekly prospect capacity by scenario
SCENARIO_CAPACITY_WEEKLY = [50, 450, 900]


# ---------------------------------------------------------------------------
# Sheet 1: Inputs
# ---------------------------------------------------------------------------
def build_inputs(wb):
    ws = wb.active
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    for c_letter in "CDEFG":
        ws.column_dimensions[c_letter].width = 16
    ws.column_dimensions["H"].width = 32

    title_rows(ws, 8)

    # ── NEWPORT BUSINESS INPUTS (yellow highlight) ────────────
    banner(ws, 5, 2, 8, "NEWPORT BUSINESS INPUTS")

    sc(ws, 6, 2, "Newport provides these values (yellow = Newport input)",
       font=NOTE_FONT)
    for c in range(3, 9):
        sc(ws, 6, c, None)
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=8)

    # Deal size headers
    sc(ws, 7, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 7, 3, "Value", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(4, 8):
        sc(ws, 7, c, None, fill=LGRAY_FILL)
    sc(ws, 7, 8, "Notes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    # Row 8: Avg deal size per segment
    seg_labels = ["Seg A", "Seg B", "Seg C", "Seg D", "Seg E"]
    sc(ws, 8, 2, "Avg Deal Size by Segment", font=LABEL_BOLD, fill=YELLOW_FILL)
    for i, seg in enumerate(SEGMENTS):
        sc(ws, 8, 3 + i, seg["deal_size"], font=INPUT_FONT,
           fill=YELLOW_FILL, fmt=FMT_CUR, align=CENTER)
    sc(ws, 8, 8, "Seg A / B / C / D / E (Newport provides actuals)",
       font=NOTE_FONT, fill=YELLOW_FILL, align=LEFT_WRAP)

    # Row 9: Segment labels
    sc(ws, 9, 2, None)
    for i, s in enumerate(seg_labels):
        sc(ws, 9, 3 + i, s, font=NOTE_FONT, align=CENTER)
    sc(ws, 9, 8, None)

    # Row 10: Gross margin
    sc(ws, 10, 2, "Current Gross Margin", font=LABEL_FONT, fill=YELLOW_FILL)
    sc(ws, 10, 3, 0.22, font=INPUT_FONT, fill=YELLOW_FILL,
       fmt=FMT_PCT, align=CENTER)
    for c in range(4, 8):
        sc(ws, 10, c, None, fill=YELLOW_FILL)
    ws.merge_cells(start_row=10, start_column=3, end_row=10, end_column=7)
    sc(ws, 10, 8, "Blended margin on commercial deals",
       font=NOTE_FONT, fill=YELLOW_FILL, align=LEFT_WRAP)

    # Row 11: Close rate on warm intros
    sc(ws, 11, 2, "Close Rate on Warm Introductions",
       font=LABEL_FONT, fill=YELLOW_FILL)
    sc(ws, 11, 3, 0.25, font=INPUT_FONT, fill=YELLOW_FILL,
       fmt=FMT_PCT, align=CENTER)
    for c in range(4, 8):
        sc(ws, 11, c, None, fill=YELLOW_FILL)
    ws.merge_cells(start_row=11, start_column=3, end_row=11, end_column=7)
    sc(ws, 11, 8, "Baseline for comparison with cold outreach",
       font=NOTE_FONT, fill=YELLOW_FILL, align=LEFT_WRAP)

    # Row 12: Sales reps
    sc(ws, 12, 2, "Number of Sales Reps", font=LABEL_FONT, fill=YELLOW_FILL)
    sc(ws, 12, 3, 2, font=INPUT_FONT, fill=YELLOW_FILL,
       fmt=FMT_NUM, align=CENTER)
    for c in range(4, 8):
        sc(ws, 12, c, None, fill=YELLOW_FILL)
    ws.merge_cells(start_row=12, start_column=3, end_row=12, end_column=7)
    sc(ws, 12, 8, "Reps handling inbound meetings from SDR",
       font=NOTE_FONT, fill=YELLOW_FILL, align=LEFT_WRAP)

    # Row 13: Monthly meeting capacity per rep
    sc(ws, 13, 2, "Monthly Meeting Capacity per Rep",
       font=LABEL_FONT, fill=YELLOW_FILL)
    sc(ws, 13, 3, 20, font=INPUT_FONT, fill=YELLOW_FILL,
       fmt=FMT_NUM, align=CENTER)
    for c in range(4, 8):
        sc(ws, 13, c, None, fill=YELLOW_FILL)
    ws.merge_cells(start_row=13, start_column=3, end_row=13, end_column=7)
    sc(ws, 13, 8, "Max new meetings one rep can handle per month",
       font=NOTE_FONT, fill=YELLOW_FILL, align=LEFT_WRAP)

    # ── OUTREACH PARAMETERS (blue = editable) ─────────────────
    banner(ws, 15, 2, 8, "OUTREACH PARAMETERS")

    sc(ws, 16, 2, "Blue values are editable assumptions", font=NOTE_FONT)
    for c in range(3, 9):
        sc(ws, 16, c, None)
    ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=8)

    sc(ws, 17, 2, "Parameter", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 17, 3, "Value", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(4, 8):
        sc(ws, 17, c, None, fill=LGRAY_FILL)
    sc(ws, 17, 8, "Notes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    outreach = [
        (18, "Emails per Day per Mailbox", 30, FMT_NUM,
         "Industry best practice: 30-50/day for deliverability"),
        (19, "Number of Mailboxes", 3, FMT_NUM,
         "Each mailbox = separate sending domain"),
        (20, "Email Open Rate", 0.22, FMT_PCT,
         "B2B benchmark: 18-25%"),
        (21, "Reply Rate", 0.035, FMT_PCT,
         "Cold email benchmark: 2-5%"),
        (22, "Reply-to-Meeting Conversion", 0.35, FMT_PCT,
         "35% of positive replies book a meeting"),
        (23, "Meeting-to-Deal Close Rate", 0.18, FMT_PCT,
         "Cold outreach close rate (lower than warm)"),
        (24, "Avg Sales Cycle (days)", 45, FMT_NUM,
         "Days from first meeting to closed deal"),
    ]

    for r, label, val, fmt, note in outreach:
        sc(ws, r, 2, label, font=LABEL_FONT)
        sc(ws, r, 3, val, font=INPUT_FONT, fmt=fmt, align=CENTER)
        for c in range(4, 8):
            sc(ws, r, c, None)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        sc(ws, r, 8, note, font=NOTE_FONT, align=LEFT_WRAP)

    # ── TOOL COSTS (3 scenarios in columns) ───────────────────
    banner(ws, 26, 2, 8, "TOOL COSTS BY SCENARIO")

    sc(ws, 27, 2, "Tool", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 27, 3, "Free", font=GREEN_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 27, 4, "Moderate", font=BLUE_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 27, 5, "Aggressive", font=ORANGE_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(6, 8):
        sc(ws, 27, c, None, fill=LGRAY_FILL)
    sc(ws, 27, 8, "Notes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    tool_notes = [
        "Free search; paid = more reveal credits",
        "Email automation platform",
        "Data enrichment & workflow",
        "SMS follow-up (usage-based)",
        "AI voice outreach (usage-based)",
    ]

    for i, (tool, free, mod, agg) in enumerate(TOOL_COSTS):
        r = 28 + i
        f = zebra(r)
        sc(ws, r, 2, tool, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, free, font=INPUT_FONT, fill=f, fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 4, mod, font=INPUT_FONT, fill=f, fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 5, agg, font=INPUT_FONT, fill=f, fmt=FMT_CUR, align=CENTER)
        for c in range(6, 8):
            sc(ws, r, c, None, fill=f)
        sc(ws, r, 8, tool_notes[i], font=NOTE_FONT, fill=f, align=LEFT_WRAP)

    # Total row (row 33)
    sc(ws, 33, 2, "TOTAL ANNUAL", font=TOTAL_FONT, fill=YELLOW_FILL)
    sc(ws, 33, 3, "=SUM(C28:C32)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_CUR, align=CENTER)
    sc(ws, 33, 4, "=SUM(D28:D32)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_CUR, align=CENTER)
    sc(ws, 33, 5, "=SUM(E28:E32)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_CUR, align=CENTER)
    for c in range(6, 8):
        sc(ws, 33, c, None, fill=YELLOW_FILL)
    sc(ws, 33, 8, None, fill=YELLOW_FILL)

    # ── CAPACITY BY SCENARIO ──────────────────────────────────
    banner(ws, 35, 2, 8, "CAPACITY BY SCENARIO")

    sc(ws, 36, 2, "Scenario", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 36, 3, "Prospects/Week", font=LABEL_BOLD,
       fill=LGRAY_FILL, align=CENTER)
    sc(ws, 36, 4, "Prospects/Month", font=LABEL_BOLD,
       fill=LGRAY_FILL, align=CENTER)
    sc(ws, 36, 5, "Method", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(6, 8):
        sc(ws, 36, c, None, fill=LGRAY_FILL)
    sc(ws, 36, 8, None, fill=LGRAY_FILL)

    cap_data = [
        (37, "Free", 50, "=C37*4.33", GREEN_FONT,
         "Manual LinkedIn + Apollo search"),
        (38, "Moderate", 450, "=C38*4.33", BLUE_FONT,
         "Email automation (Instantly)"),
        (39, "Aggressive", 900, "=C39*4.33", ORANGE_FONT,
         "Multi-channel (email + SMS + voice)"),
    ]

    for r, label, weekly, monthly_formula, lfont, method in cap_data:
        f = zebra(r)
        sc(ws, r, 2, label, font=lfont, fill=f)
        sc(ws, r, 3, weekly, font=INPUT_FONT, fill=f, fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 4, monthly_formula, font=FORMULA_FONT, fill=f,
           fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 5, method, font=LABEL_FONT, fill=f, align=LEFT_WRAP)
        for c in range(6, 8):
            sc(ws, r, c, None, fill=f)
        sc(ws, r, 8, None, fill=f)

    # ── WEIGHTED AVG DEAL SIZE (computed) ─────────────────────
    banner(ws, 41, 2, 8, "WEIGHTED AVERAGE DEAL SIZE")

    sc(ws, 42, 2, "Computed from segment deal sizes (equal weight)",
       font=NOTE_FONT)
    for c in range(3, 9):
        sc(ws, 42, c, None)
    ws.merge_cells(start_row=42, start_column=2, end_row=42, end_column=8)

    sc(ws, 43, 2, "Weighted Avg Deal Size", font=LABEL_BOLD, fill=YELLOW_FILL)
    sc(ws, 43, 3, "=AVERAGE(C8:G8)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_CUR, align=CENTER)
    for c in range(4, 8):
        sc(ws, 43, c, None, fill=YELLOW_FILL)
    ws.merge_cells(start_row=43, start_column=3, end_row=43, end_column=7)
    sc(ws, 43, 8, "Average across 5 segments", font=NOTE_FONT,
       fill=YELLOW_FILL, align=LEFT_WRAP)

    return ws


# ---------------------------------------------------------------------------
# Sheet 2: ICP Segments
# ---------------------------------------------------------------------------
def build_icp_segments(wb):
    ws = wb.create_sheet("ICP Segments")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    for c_letter in "CDEFG":
        ws.column_dimensions[c_letter].width = 20

    title_rows(ws, 7)

    banner(ws, 5, 2, 7, "ICP SEGMENT ECONOMICS")

    # Header row with segment names
    sc(ws, 6, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for i, seg in enumerate(SEGMENTS):
        sc(ws, 6, 3 + i, f"Seg {seg['code']}: {seg['name']}",
           font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    # Metric rows
    metrics = [
        ("Pipeline Type", "pipeline", None, LABEL_FONT, None),
        ("Geography", "geo", None, LABEL_FONT, None),
        ("Est. Addressable Companies", "companies", FMT_NUM, FORMULA_FONT, None),
        ("Apollo Coverage", "apollo_cov", FMT_PCT, FORMULA_FONT, None),
    ]

    r = 7
    for label, key, fmt, font, _ in metrics:
        f = zebra(r)
        sc(ws, r, 2, label, font=LABEL_FONT, fill=f)
        for i, seg in enumerate(SEGMENTS):
            sc(ws, r, 3 + i, seg[key], font=font, fill=f,
               fmt=fmt, align=CENTER)
        r += 1

    # Row 11: Contacts per Company
    sc(ws, 11, 2, "Contacts per Company", font=LABEL_FONT, fill=zebra(11))
    for i in range(5):
        sc(ws, 11, 3 + i, 3, font=INPUT_FONT, fill=zebra(11),
           fmt=FMT_NUM, align=CENTER)

    # Row 12: Reachable Contacts (formula: companies * coverage * contacts/co)
    f12 = zebra(12)
    sc(ws, 12, 2, "Reachable Contacts", font=LABEL_BOLD, fill=f12)
    # companies in row 9, coverage in row 10, contacts/co in row 11
    for i in range(5):
        col = get_column_letter(3 + i)
        sc(ws, 12, 3 + i,
           f"=ROUND({col}9*{col}10*{col}11,0)",
           font=FORMULA_FONT, fill=f12, fmt=FMT_NUM, align=CENTER)

    # Row 13: blank spacer
    r = 13
    for c in range(2, 8):
        sc(ws, r, c, None)

    # Row 14: Avg Deal Size (references Inputs sheet)
    f14 = zebra(14)
    sc(ws, 14, 2, "Avg Deal Size", font=LABEL_BOLD, fill=f14)
    seg_cols_inputs = ["C", "D", "E", "F", "G"]
    for i in range(5):
        sc(ws, 14, 3 + i, f"=Inputs!{seg_cols_inputs[i]}8",
           font=INPUT_FONT, fill=f14, fmt=FMT_CUR, align=CENTER)

    # Row 15: Gross Margin (references Inputs)
    f15 = zebra(15)
    sc(ws, 15, 2, "Gross Margin", font=LABEL_FONT, fill=f15)
    for i in range(5):
        sc(ws, 15, 3 + i, "=Inputs!C10",
           font=FORMULA_FONT, fill=f15, fmt=FMT_PCT, align=CENTER)

    # Row 16: Est. Annual Revenue per Deal
    f16 = zebra(16)
    sc(ws, 16, 2, "Est. Annual Revenue per Deal", font=LABEL_FONT, fill=f16)
    for i in range(5):
        col = get_column_letter(3 + i)
        sc(ws, 16, 3 + i, f"={col}14",
           font=FORMULA_FONT, fill=f16, fmt=FMT_CUR, align=CENTER)

    # Row 17: Gross Profit per Deal
    f17 = zebra(17)
    sc(ws, 17, 2, "Gross Profit per Deal", font=LABEL_FONT, fill=f17)
    for i in range(5):
        col = get_column_letter(3 + i)
        sc(ws, 17, 3 + i, f"={col}16*{col}15",
           font=FORMULA_FONT, fill=f17, fmt=FMT_CUR, align=CENTER)

    # Row 18: blank spacer
    for c in range(2, 8):
        sc(ws, 18, c, None)

    # Row 19: Segment Priority
    f19 = zebra(19)
    sc(ws, 19, 2, "Segment Priority", font=LABEL_BOLD, fill=f19)
    for i, seg in enumerate(SEGMENTS):
        pri = seg["priority"]
        pri_font = GREEN_BOLD if pri == 1 else (BLUE_FONT if pri == 2 else LABEL_FONT)
        sc(ws, 19, 3 + i, pri, font=pri_font, fill=f19,
           fmt=FMT_NUM, align=CENTER)

    # Row 20: Notes
    f20 = zebra(20)
    sc(ws, 20, 2, "Notes", font=LABEL_FONT, fill=f20)
    for i, seg in enumerate(SEGMENTS):
        sc(ws, 20, 3 + i, seg["notes"], font=NOTE_FONT, fill=f20,
           align=LEFT_WRAP)

    # ── SEGMENT TAM SUMMARY ──────────────────────────────────
    banner(ws, 22, 2, 7, "SEGMENT TAM SUMMARY")

    sc(ws, 23, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for i, seg in enumerate(SEGMENTS):
        sc(ws, 23, 3 + i, f"Seg {seg['code']}",
           font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    # Row 24: TAM = companies * deal_size
    f24 = zebra(24)
    sc(ws, 24, 2, "Segment TAM (companies x deal)", font=LABEL_FONT, fill=f24)
    for i in range(5):
        col = get_column_letter(3 + i)
        sc(ws, 24, 3 + i, f"={col}9*{col}14",
           font=FORMULA_FONT, fill=f24, fmt=FMT_CUR, align=CENTER)

    # Row 25: Reachable TAM = reachable contacts / contacts_per_co * deal_size
    f25 = zebra(25)
    sc(ws, 25, 2, "Reachable TAM (Apollo-filtered)",
       font=LABEL_FONT, fill=f25)
    for i in range(5):
        col = get_column_letter(3 + i)
        sc(ws, 25, 3 + i, f"=ROUND({col}12/{col}11,0)*{col}14",
           font=FORMULA_FONT, fill=f25, fmt=FMT_CUR, align=CENTER)

    return ws


# ---------------------------------------------------------------------------
# Sheet 3: Funnel Model (12-month, 3 scenarios)
# ---------------------------------------------------------------------------
def build_funnel_model(wb):
    ws = wb.create_sheet("Funnel Model")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28

    # 3 scenario blocks of 14 cols each (M1-M12 + spacer + Total)
    # Scenario layout: each takes columns for label + M1..M12 + Total
    # We'll place scenarios vertically (stacked) for readability

    # Column widths: B=label, C-N = months, O = total
    for c in range(3, 16):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.column_dimensions[get_column_letter(15)].width = 14  # Total col

    title_rows(ws, 15)

    # Sales cycle lag in months (45 days ~ 2 months)
    # Deals closed = meetings from 2 months prior * close_rate

    # Build 3 scenario blocks stacked vertically
    scenario_configs = [
        ("FREE SCENARIO", GREEN_BOLD, 37),     # capacity row in Inputs
        ("MODERATE SCENARIO", BLUE_BOLD, 38),
        ("AGGRESSIVE SCENARIO", ORANGE_BOLD, 39),
    ]

    # Scenario tool cost rows in Inputs: C33=Free, D33=Mod, E33=Agg
    cost_cols = ["C", "D", "E"]

    block_start_rows = []
    current_row = 5

    for s_idx, (scenario_name, s_font, cap_row) in enumerate(scenario_configs):
        start_r = current_row
        block_start_rows.append(start_r)

        # Banner
        banner(ws, start_r, 2, 15, scenario_name)

        # Month headers
        hr = start_r + 1
        sc(ws, hr, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL)
        for m in range(1, 13):
            sc(ws, hr, 2 + m, f"M{m}", font=LABEL_BOLD,
               fill=LGRAY_FILL, align=CENTER)
        sc(ws, hr, 15, "Total", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

        # Row offsets from start_r:
        #  +2 = Prospects Contacted
        #  +3 = Opens
        #  +4 = Replies
        #  +5 = Meetings Booked
        #  +6 = Meetings (capacity-capped)
        #  +7 = Deals Closed (lagged)
        #  +8 = Revenue
        #  +9 = Gross Profit
        # +10 = Tool Costs
        # +11 = Net Contribution
        # +12 = Cumulative Revenue

        r_contacted = start_r + 2
        r_opens = start_r + 3
        r_replies = start_r + 4
        r_meetings = start_r + 5
        r_meetings_cap = start_r + 6
        r_deals = start_r + 7
        r_revenue = start_r + 8
        r_gp = start_r + 9
        r_tool = start_r + 10
        r_net = start_r + 11
        r_cumrev = start_r + 12

        # -- Prospects Contacted --
        sc(ws, r_contacted, 2, "Prospects Contacted",
           font=LABEL_FONT, fill=zebra(r_contacted))
        for m in range(1, 13):
            mc = 2 + m
            sc(ws, r_contacted, mc,
               f"=ROUND(Inputs!D{cap_row},0)",
               font=FORMULA_FONT, fill=zebra(r_contacted),
               fmt=FMT_NUM, align=CENTER)
        sc(ws, r_contacted, 15,
           f"=SUM(C{r_contacted}:N{r_contacted})",
           font=TOTAL_FONT, fill=zebra(r_contacted),
           fmt=FMT_NUM, align=CENTER)

        # -- Opens --
        sc(ws, r_opens, 2, "Opens", font=LABEL_FONT, fill=zebra(r_opens))
        for m in range(1, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            sc(ws, r_opens, mc,
               f"=ROUND({mcl}{r_contacted}*Inputs!C20,0)",
               font=FORMULA_FONT, fill=zebra(r_opens),
               fmt=FMT_NUM, align=CENTER)
        sc(ws, r_opens, 15,
           f"=SUM(C{r_opens}:N{r_opens})",
           font=TOTAL_FONT, fill=zebra(r_opens),
           fmt=FMT_NUM, align=CENTER)

        # -- Replies --
        sc(ws, r_replies, 2, "Replies", font=LABEL_FONT, fill=zebra(r_replies))
        for m in range(1, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            sc(ws, r_replies, mc,
               f"=ROUND({mcl}{r_contacted}*Inputs!C21,0)",
               font=FORMULA_FONT, fill=zebra(r_replies),
               fmt=FMT_NUM, align=CENTER)
        sc(ws, r_replies, 15,
           f"=SUM(C{r_replies}:N{r_replies})",
           font=TOTAL_FONT, fill=zebra(r_replies),
           fmt=FMT_NUM, align=CENTER)

        # -- Meetings Booked (raw) --
        sc(ws, r_meetings, 2, "Meetings Booked (raw)",
           font=LABEL_FONT, fill=zebra(r_meetings))
        for m in range(1, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            sc(ws, r_meetings, mc,
               f"=ROUND({mcl}{r_replies}*Inputs!C22,0)",
               font=FORMULA_FONT, fill=zebra(r_meetings),
               fmt=FMT_NUM, align=CENTER)
        sc(ws, r_meetings, 15,
           f"=SUM(C{r_meetings}:N{r_meetings})",
           font=TOTAL_FONT, fill=zebra(r_meetings),
           fmt=FMT_NUM, align=CENTER)

        # -- Meetings (capacity-capped) --
        sc(ws, r_meetings_cap, 2, "Meetings (capacity-capped)",
           font=LABEL_BOLD, fill=zebra(r_meetings_cap))
        for m in range(1, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            sc(ws, r_meetings_cap, mc,
               f"=MIN({mcl}{r_meetings},Inputs!C12*Inputs!C13)",
               font=FORMULA_FONT, fill=zebra(r_meetings_cap),
               fmt=FMT_NUM, align=CENTER)
        sc(ws, r_meetings_cap, 15,
           f"=SUM(C{r_meetings_cap}:N{r_meetings_cap})",
           font=TOTAL_FONT, fill=zebra(r_meetings_cap),
           fmt=FMT_NUM, align=CENTER)

        # -- Deals Closed (lagged by ~2 months for 45-day sales cycle) --
        sc(ws, r_deals, 2, "Deals Closed",
           font=LABEL_BOLD, fill=zebra(r_deals))
        for m in range(1, 13):
            mc = 2 + m
            if m <= 2:
                # No deals close in first 2 months (45-day sales cycle)
                sc(ws, r_deals, mc, 0, font=FORMULA_FONT,
                   fill=zebra(r_deals), fmt=FMT_NUM, align=CENTER)
            else:
                # Meetings from 2 months ago * close rate
                lagged_col = get_column_letter(mc - 2)
                sc(ws, r_deals, mc,
                   f"=ROUND({lagged_col}{r_meetings_cap}*Inputs!C23,0)",
                   font=FORMULA_FONT, fill=zebra(r_deals),
                   fmt=FMT_NUM, align=CENTER)
        sc(ws, r_deals, 15,
           f"=SUM(C{r_deals}:N{r_deals})",
           font=TOTAL_FONT, fill=zebra(r_deals),
           fmt=FMT_NUM, align=CENTER)

        # -- Revenue --
        sc(ws, r_revenue, 2, "Revenue",
           font=LABEL_FONT, fill=zebra(r_revenue))
        for m in range(1, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            sc(ws, r_revenue, mc,
               f"={mcl}{r_deals}*Inputs!C43",
               font=FORMULA_FONT, fill=zebra(r_revenue),
               fmt=FMT_CUR, align=CENTER)
        sc(ws, r_revenue, 15,
           f"=SUM(C{r_revenue}:N{r_revenue})",
           font=TOTAL_FONT, fill=zebra(r_revenue),
           fmt=FMT_CUR, align=CENTER)

        # -- Gross Profit --
        sc(ws, r_gp, 2, "Gross Profit",
           font=LABEL_FONT, fill=zebra(r_gp))
        for m in range(1, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            sc(ws, r_gp, mc,
               f"={mcl}{r_revenue}*Inputs!C10",
               font=FORMULA_FONT, fill=zebra(r_gp),
               fmt=FMT_CUR, align=CENTER)
        sc(ws, r_gp, 15,
           f"=SUM(C{r_gp}:N{r_gp})",
           font=TOTAL_FONT, fill=zebra(r_gp),
           fmt=FMT_CUR, align=CENTER)

        # -- Tool Costs (monthly = annual / 12) --
        sc(ws, r_tool, 2, "Tool Costs (monthly)",
           font=LABEL_FONT, fill=zebra(r_tool))
        cost_col = cost_cols[s_idx]
        for m in range(1, 13):
            mc = 2 + m
            sc(ws, r_tool, mc,
               f"=ROUND(Inputs!{cost_col}33/12,0)",
               font=FORMULA_FONT, fill=zebra(r_tool),
               fmt=FMT_CUR, align=CENTER)
        sc(ws, r_tool, 15,
           f"=SUM(C{r_tool}:N{r_tool})",
           font=TOTAL_FONT, fill=zebra(r_tool),
           fmt=FMT_CUR, align=CENTER)

        # -- Net Contribution --
        sc(ws, r_net, 2, "NET CONTRIBUTION",
           font=TOTAL_FONT, fill=YELLOW_FILL)
        for m in range(1, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            sc(ws, r_net, mc,
               f"={mcl}{r_gp}-{mcl}{r_tool}",
               font=TOTAL_FONT, fill=YELLOW_FILL,
               fmt=FMT_CUR, align=CENTER)
        sc(ws, r_net, 15,
           f"=SUM(C{r_net}:N{r_net})",
           font=TOTAL_FONT, fill=YELLOW_FILL,
           fmt=FMT_CUR, align=CENTER)

        # -- Cumulative Revenue --
        sc(ws, r_cumrev, 2, "Cumulative Revenue",
           font=LABEL_FONT, fill=zebra(r_cumrev))
        sc(ws, r_cumrev, 3,
           f"=C{r_revenue}",
           font=FORMULA_FONT, fill=zebra(r_cumrev),
           fmt=FMT_CUR, align=CENTER)
        for m in range(2, 13):
            mc = 2 + m
            mcl = get_column_letter(mc)
            prev_cl = get_column_letter(mc - 1)
            sc(ws, r_cumrev, mc,
               f"={prev_cl}{r_cumrev}+{mcl}{r_revenue}",
               font=FORMULA_FONT, fill=zebra(r_cumrev),
               fmt=FMT_CUR, align=CENTER)
        sc(ws, r_cumrev, 15,
           f"=N{r_cumrev}",
           font=TOTAL_FONT, fill=zebra(r_cumrev),
           fmt=FMT_CUR, align=CENTER)

        current_row = r_cumrev + 2  # gap before next scenario

    # ═══════════════════════════════════════════════════════════
    # SUMMARY TABLE (after all 3 scenario blocks)
    # ═══════════════════════════════════════════════════════════
    summary_r = current_row
    banner(ws, summary_r, 2, 15, "12-MONTH SUMMARY")

    sh = summary_r + 1
    sc(ws, sh, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, sh, 3, "Free", font=GREEN_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, sh, 4, "Moderate", font=BLUE_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, sh, 5, "Aggressive", font=ORANGE_BOLD,
       fill=LGRAY_FILL, align=CENTER)
    for c in range(6, 16):
        sc(ws, sh, c, None, fill=LGRAY_FILL)

    # References to each scenario's Total column (O = col 15)
    free_base = block_start_rows[0]
    mod_base = block_start_rows[1]
    agg_base = block_start_rows[2]

    summary_metrics = [
        ("Prospects Contacted", 2),
        ("Meetings Booked", 6),
        ("Deals Closed", 7),
        ("Revenue", 8),
        ("Gross Profit", 9),
        ("Tool Costs", 10),
        ("Net Contribution", 11),
    ]

    for i, (label, offset) in enumerate(summary_metrics):
        r = sh + 1 + i
        f = zebra(r)
        is_total = (label == "Net Contribution")
        use_fill = YELLOW_FILL if is_total else f
        use_font = TOTAL_FONT if is_total else LABEL_FONT
        fmt = FMT_CUR if offset >= 8 else FMT_NUM

        sc(ws, r, 2, label, font=use_font, fill=use_fill)
        sc(ws, r, 3,
           f"=O{free_base + offset}", font=GREEN_FONT, fill=use_fill,
           fmt=fmt, align=CENTER)
        sc(ws, r, 4,
           f"=O{mod_base + offset}", font=BLUE_FONT, fill=use_fill,
           fmt=fmt, align=CENTER)
        sc(ws, r, 5,
           f"=O{agg_base + offset}", font=ORANGE_FONT, fill=use_fill,
           fmt=fmt, align=CENTER)
        for c in range(6, 16):
            sc(ws, r, c, None, fill=use_fill)

    # ═══════════════════════════════════════════════════════════
    # CHARTS
    # ═══════════════════════════════════════════════════════════
    chart_start = sh + len(summary_metrics) + 2

    # Chart 1: Cumulative Revenue by Scenario (line chart, 3 lines)
    chart1 = LineChart()
    chart1.title = "Cumulative Revenue by Scenario (12 Months)"
    chart1.y_axis.title = "Revenue ($)"
    chart1.style = 10
    chart1.width = 24
    chart1.height = 14

    # Month headers from Free scenario row
    cats = Reference(ws, min_col=3, max_col=14,
                     min_row=block_start_rows[0] + 1)

    colors = ["548235", "2E5090", "ED7D31"]
    labels = ["Free", "Moderate", "Aggressive"]
    for s_idx, base in enumerate(block_start_rows):
        cumrev_row = base + 12
        data = Reference(ws, min_col=3, max_col=14, min_row=cumrev_row)
        chart1.add_data(data, from_rows=True, titles_from_data=False)
        chart1.series[-1].tx = SeriesLabel(v=labels[s_idx])
        chart1.series[-1].graphicalProperties.line.solidFill = colors[s_idx]
    chart1.set_categories(cats)
    ws.add_chart(chart1, f"B{chart_start}")

    # Chart 2: Monthly Meetings Booked by Scenario (bar chart)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Monthly Meetings Booked by Scenario"
    chart2.y_axis.title = "Meetings"
    chart2.style = 10
    chart2.width = 24
    chart2.height = 14

    for s_idx, base in enumerate(block_start_rows):
        meetings_row = base + 6  # capacity-capped meetings
        data = Reference(ws, min_col=3, max_col=14, min_row=meetings_row)
        chart2.add_data(data, from_rows=True, titles_from_data=False)
        chart2.series[-1].tx = SeriesLabel(v=labels[s_idx])
    chart2.set_categories(cats)
    ws.add_chart(chart2, f"B{chart_start + 16}")

    # Chart 3: 12-Month Net Contribution by Scenario (summary bar)
    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "clustered"
    chart3.title = "12-Month Net Contribution by Scenario"
    chart3.y_axis.title = "Net Contribution ($)"
    chart3.style = 10
    chart3.width = 16
    chart3.height = 12

    # Net contribution row in summary table
    net_summary_row = sh + len(summary_metrics)  # last summary row
    data3 = Reference(ws, min_col=3, max_col=5, min_row=net_summary_row)
    cats3 = Reference(ws, min_col=3, max_col=5, min_row=sh)
    chart3.add_data(data3, from_rows=True, titles_from_data=False)
    chart3.series[0].tx = SeriesLabel(v="Net Contribution")
    chart3.set_categories(cats3)
    ws.add_chart(chart3, f"B{chart_start + 32}")

    return ws


# ---------------------------------------------------------------------------
# Sheet 4: Market Analysis
# ---------------------------------------------------------------------------
def build_market_analysis(wb):
    ws = wb.create_sheet("Market Analysis")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 28

    title_rows(ws, 7)

    # ── MARKET SIZING BY SEGMENT ──────────────────────────────
    banner(ws, 5, 2, 7, "MARKET SIZING BY SEGMENT")

    sc(ws, 6, 2, "Segment", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 3, "Companies", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 4, "Reachable (Apollo)", font=LABEL_BOLD,
       fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 5, "Avg Deal", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 6, "Est. TAM", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 7, "Notes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    seg_data = [
        ("A: Enterprise Grocery", 2000, 0.65, 50000,
         "Largest deal sizes; centralized procurement"),
        ("B: Manufacturers", 3000, 0.60, 25000,
         "Volume play; partner pipeline"),
        ("C: Gov Buyers", 500, 0.40, 15000,
         "Overlap with GovCon; Apollo coverage low"),
        ("D: Corrections", 200, 0.55, 30000,
         "Niche; high barriers but sticky contracts"),
        ("E: Candy/Confectionery", 1831, 0.60, 20000,
         "NAICS 424450; Newport's specialty"),
    ]

    for i, (seg, cos, cov, deal, note) in enumerate(seg_data):
        r = 7 + i
        f = zebra(r)
        reachable = round(cos * cov)
        tam = cos * deal
        sc(ws, r, 2, seg, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, cos, font=FORMULA_FONT, fill=f, fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 4, reachable, font=FORMULA_FONT, fill=f,
           fmt=FMT_NUM, align=CENTER)
        sc(ws, r, 5, deal, font=FORMULA_FONT, fill=f,
           fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 6, tam, font=FORMULA_FONT, fill=f,
           fmt=FMT_CUR, align=CENTER)
        sc(ws, r, 7, note, font=NOTE_FONT, fill=f, align=LEFT_WRAP)

    # Total row
    sc(ws, 12, 2, "TOTAL", font=TOTAL_FONT, fill=YELLOW_FILL)
    sc(ws, 12, 3, "=SUM(C7:C11)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_NUM, align=CENTER)
    sc(ws, 12, 4, "=SUM(D7:D11)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_NUM, align=CENTER)
    sc(ws, 12, 5, None, fill=YELLOW_FILL)
    sc(ws, 12, 6, "=SUM(F7:F11)", font=TOTAL_FONT, fill=YELLOW_FILL,
       fmt=FMT_CUR, align=CENTER)
    sc(ws, 12, 7, None, fill=YELLOW_FILL)

    # TAM Pie chart
    pie = PieChart()
    pie.title = "Estimated TAM by Segment"
    pie.style = 10
    pie.width = 16
    pie.height = 12
    pie_data = Reference(ws, min_col=6, min_row=7, max_row=11)
    pie_cats = Reference(ws, min_col=2, min_row=7, max_row=11)
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_cats)
    ws.add_chart(pie, "B14")

    # ── CANDY WHOLESALE MARKET ────────────────────────────────
    banner(ws, 28, 2, 7, "CANDY WHOLESALE MARKET (NAICS 424450)")

    candy_data = [
        ("US Firms (NAICS 424450)", "1,831", "US Census Bureau"),
        ("US Establishments", "2,357", "US Census Bureau"),
        ("US Employees", "40,279", "US Census Bureau"),
        ("US Retail Confectionery Market", "$38B", "Grand View Research"),
        ("LATAM Confectionery Market", "$35-42B", "IndexBox / Mordor Intelligence"),
        ("NCA Member Retail Sales", "$54B", "National Confectioners Assoc."),
        ("Key Trade Agreement", "USMCA: $0 tariff on MX candy", "USDA FAS"),
        ("CAFTA-DR", "Free/near-free on CA candy", "USDA FAS"),
    ]

    sc(ws, 29, 2, "Metric", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 29, 3, "Value", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    for c in range(4, 7):
        sc(ws, 29, c, None, fill=LGRAY_FILL)
    sc(ws, 29, 7, "Source", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    for i, (metric, val, source) in enumerate(candy_data):
        r = 30 + i
        f = zebra(r)
        sc(ws, r, 2, metric, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, val, font=FORMULA_FONT, fill=f, align=CENTER)
        for c in range(4, 7):
            sc(ws, r, c, None, fill=f)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        sc(ws, r, 7, source, font=NOTE_FONT, fill=f, align=LEFT_WRAP)

    # ── TOOL COMPARISON MATRIX ────────────────────────────────
    banner(ws, 39, 2, 7, "TOOL COMPARISON MATRIX")

    sc(ws, 40, 2, "Tool", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 40, 3, "Purpose", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 40, 4, "Monthly Cost", font=LABEL_BOLD,
       fill=LGRAY_FILL, align=CENTER)
    sc(ws, 40, 5, "Annual Cost", font=LABEL_BOLD,
       fill=LGRAY_FILL, align=CENTER)
    sc(ws, 40, 6, "Scenario", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 40, 7, "Value Proposition", font=LABEL_BOLD,
       fill=LGRAY_FILL, align=CENTER)

    tools = [
        ("Apollo.io (Basic)", "Contact search & enrichment",
         "$49", "$588", "Moderate",
         "Unlimited search; 900 credits/mo for email reveals"),
        ("Apollo.io (Pro)", "Advanced enrichment + intent",
         "$119", "$1,428", "Aggressive",
         "Buyer intent signals, advanced filters, API access"),
        ("Instantly (Growth)", "Email automation & warmup",
         "$37", "$444", "Moderate",
         "1,000 active contacts; unlimited emails; warmup"),
        ("Instantly (Hypergrowth)", "Multi-account email",
         "$97", "$1,164", "Aggressive",
         "25,000 contacts; 25 email accounts; A/B testing"),
        ("Clay", "Data enrichment workflows",
         "$390", "$4,680", "Aggressive",
         "Waterfall enrichment; 75+ data providers"),
        ("Twilio SMS", "SMS follow-up", "~$50", "~$600", "Aggressive",
         "Multi-channel touchpoint; high open rate"),
        ("Retell AI Voice", "AI voice outreach", "~$100", "~$1,200",
         "Aggressive", "AI voice agent for meeting scheduling"),
    ]

    for i, (tool, purpose, monthly, annual, scenario, value) in enumerate(tools):
        r = 41 + i
        f = zebra(r)
        sc(ws, r, 2, tool, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, purpose, font=LABEL_FONT, fill=f, align=LEFT_WRAP)
        sc(ws, r, 4, monthly, font=INPUT_FONT, fill=f, align=CENTER)
        sc(ws, r, 5, annual, font=INPUT_FONT, fill=f, align=CENTER)
        sc(ws, r, 6, scenario, font=LABEL_FONT, fill=f, align=CENTER)
        sc(ws, r, 7, value, font=NOTE_FONT, fill=f, align=LEFT_WRAP)

    # ── APOLLO COVERAGE BY REGION ─────────────────────────────
    banner(ws, 49, 2, 7, "APOLLO COVERAGE BY REGION")

    sc(ws, 50, 2, "Region", font=LABEL_BOLD, fill=LGRAY_FILL)
    sc(ws, 50, 3, "Est. Coverage", font=LABEL_BOLD,
       fill=LGRAY_FILL, align=CENTER)
    for c in range(4, 7):
        sc(ws, 50, c, None, fill=LGRAY_FILL)
    sc(ws, 50, 7, "Notes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    regions = [
        ("United States", "60-70%", "Strong coverage; major distributors well-indexed"),
        ("Mexico", "30-50%", "Decent for larger companies; supplement with SIEM registry"),
        ("Central America", "10-25%", "Smaller economies; supplement with trade directories"),
        ("South America (BR/CO/AR)", "25-40%",
         "Larger markets have better coverage"),
        ("Caribbean", "5-15%", "Small market; many informal businesses"),
        ("UK / EU", "50-65%", "Good for enterprise; varies by country"),
    ]

    for i, (region, cov, note) in enumerate(regions):
        r = 51 + i
        f = zebra(r)
        sc(ws, r, 2, region, font=LABEL_FONT, fill=f)
        sc(ws, r, 3, cov, font=FORMULA_FONT, fill=f, align=CENTER)
        for c in range(4, 7):
            sc(ws, r, c, None, fill=f)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        sc(ws, r, 7, note, font=NOTE_FONT, fill=f, align=LEFT_WRAP)

    # Sources
    sc(ws, 58, 2,
       "Sources: US Census Bureau (NAICS 424450), Grand View Research, "
       "Mordor Intelligence, IndexBox, USDA FAS, NCA, Apollo.io coverage "
       "estimates, specs/09-INTEGRATIONS.md. Feb 2026.",
       font=NOTE_FONT)
    for c in range(3, 8):
        sc(ws, 58, c, None)
    ws.merge_cells(start_row=58, start_column=2, end_row=58, end_column=7)

    return ws


# ---------------------------------------------------------------------------
# Sheet 5: Key Questions
# ---------------------------------------------------------------------------
def build_key_questions(wb):
    ws = wb.create_sheet("Key Questions")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 32

    title_rows(ws, 5)

    banner(ws, 5, 2, 5, "CRITICAL DECISIONS THAT CALIBRATE THE MODEL")

    sc(ws, 6, 2, "Priority", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 3, "Question", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 4, "If Yes", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)
    sc(ws, 6, 5, "If No", font=LABEL_BOLD, fill=LGRAY_FILL, align=CENTER)

    questions = [
        ("1 HIGHEST", RED_BOLD,
         "What are actual deal sizes by customer type?",
         "Calibrates entire model. Current values are placeholders.",
         "Model uses estimates; directionally right but imprecise"),
        ("2 HIGHEST", RED_BOLD,
         "What is current gross margin on wholesale deals?",
         "Determines profitability at every scenario level.",
         "22% assumption may over- or under-state returns"),
        ("3 HIGH", ORANGE_BOLD,
         "How many sales reps will handle inbound meetings?",
         "Capacity constraint. More reps = more throughput.",
         "2 reps cap at 40 meetings/mo; may bottleneck Aggressive"),
        ("4 HIGH", ORANGE_BOLD,
         "Which segments to prioritize first?",
         "Focus resources on 1-2 segments for faster traction.",
         "Spreading across all 5 dilutes effort and slows results"),
        ("5 HIGH", ORANGE_BOLD,
         "Willing to invest in Instantly email automation?",
         "Unlocks Moderate scenario (450/wk vs 50/wk).",
         "Free scenario only; manual LinkedIn outreach"),
        ("6 MEDIUM", BLUE_BOLD,
         "Current close rate on warm introductions?",
         "Baseline calibration. Cold outreach typically 40-60% lower.",
         "We use 18% default for cold; adjust if warm rate known"),
        ("7 MEDIUM", BLUE_BOLD,
         "Interest in LATAM candy import/export?",
         "Expands Seg E scope. USMCA = $0 tariff from Mexico.",
         "Focus US-only Seg E; still 1,831 firms to target"),
        ("8 MEDIUM", BLUE_BOLD,
         "Existing CRM or sales tracking system?",
         "Integrate SDR pipeline into existing workflow.",
         "Recommend HubSpot Free CRM or Google Sheets tracker"),
        ("9 INFO", NOTE_FONT,
         "What does a typical sales cycle look like today?",
         "Calibrates 45-day default. Shorter = faster revenue.",
         "We keep 45-day assumption; conservative but standard"),
        ("10 INFO", NOTE_FONT,
         "Any existing outbound prospecting?",
         "Baseline comparison. SDR adds to existing effort.",
         "Greenfield — all SDR revenue is incremental"),
    ]

    for i, (priority, pri_font, question, if_yes, if_no) in enumerate(questions):
        r = 7 + i
        f = zebra(r)
        ws.row_dimensions[r].height = 40
        sc(ws, r, 2, priority, font=pri_font, fill=f, align=CENTER)
        sc(ws, r, 3, question, font=LABEL_FONT, fill=f, align=LEFT_WRAP)
        sc(ws, r, 4, if_yes, font=GREEN_FONT, fill=f, align=LEFT_WRAP)
        sc(ws, r, 5, if_no, font=RED_FONT, fill=f, align=LEFT_WRAP)

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Build Newport Commercial SDR Financial Model"
    )
    parser.add_argument("--output", help="Output path (default: same directory)")
    args = parser.parse_args()

    output_dir = Path(__file__).resolve().parent
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = output_dir / "newport-commercial-sdr-model.xlsx"

    print("=" * 60)
    print("Newport Commercial SDR Financial Model Builder")
    print("=" * 60)

    wb = Workbook()

    print("[1/5] Building Inputs sheet...")
    build_inputs(wb)

    print("[2/5] Building ICP Segments sheet...")
    build_icp_segments(wb)

    print("[3/5] Building Funnel Model sheet (with 3 charts)...")
    build_funnel_model(wb)

    print("[4/5] Building Market Analysis sheet (with 1 chart)...")
    build_market_analysis(wb)

    print("[5/5] Building Key Questions sheet...")
    build_key_questions(wb)

    wb.save(str(output_path))
    print()
    print(f"Saved: {output_path}")
    print(f"File size: {output_path.stat().st_size:,} bytes")
    print()

    print("Verification:")
    print(f"  Sheets: {wb.sheetnames}")
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"  {ws_name}: {ws.max_row} rows x {ws.max_column} cols")

    print()
    print("Design: Blue text = editable inputs | Yellow = Newport-provided")
    print("Formulas reference Inputs sheet — change any input, everything recalculates")
    print("3 scenarios: Free ($0) | Moderate ($1,032/yr) | Aggressive ($9,072/yr)")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
