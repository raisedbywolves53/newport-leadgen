"""Generate Newport Wholesalers Government Contract Opportunity Dashboard.

Creates an 8-tab Excel workbook with:
  Tab 1: Active Opportunities Pipeline
  Tab 2: Expiring Contracts (Rebid Pipeline)
  Tab 3: Sources Sought Tracker
  Tab 4: Decision Maker Tracker
  Tab 5: Past Performance Portfolio
  Tab 6: Scoring Model
  Tab 7: Summary Dashboard
  Tab 8: Conference & Event Calendar

Usage:
    python dashboard/generate_dashboard.py
    python dashboard/generate_dashboard.py --contacts data/contacts/fl_decision_makers.csv
    python dashboard/generate_dashboard.py --market-data data/final/govt_market_size_by_agency_20260220_2132.csv
"""

import argparse
import csv
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_ROOT / "dashboard"

# --- Style Constants ---
NAVY = "1B2A4A"
GOLD = "D4A843"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
RED_BG = "FFCCCC"
YELLOW_BG = "FFFFCC"
GREEN_BG = "CCFFCC"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
CURRENCY_FMT = '"$"#,##0'
CURRENCY_FMT_K = '"$"#,##0"K"'
DATE_FMT = "MM/DD/YYYY"
THIN_BORDER = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)

HIGH_FILL = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
MED_FILL = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
LOW_FILL = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
ALT_ROW = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


def style_header_row(ws, num_cols, row=1):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 35
    ws.auto_filter.ref = f"A{row}:{get_column_letter(num_cols)}{row}"


def style_data_rows(ws, num_cols, start_row=2, end_row=None):
    if end_row is None:
        end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW


def auto_width(ws, num_cols, max_width=40, min_width=12):
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in range(1, min(ws.max_row + 1, 50)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


# --- Tab 1: Active Opportunities Pipeline ---
def build_opportunities_tab(wb, sample_data=None):
    ws = wb.create_sheet("Opportunities Pipeline")
    headers = [
        "Opportunity ID", "Title", "Agency", "Level",
        "Type", "Est. Value", "Response Deadline", "Days Until Deadline",
        "Priority Score", "Priority Tier", "Status",
        "Set-Aside Status", "Key Requirements",
        "Decision Maker", "Relationship Status", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    sample = sample_data or [
        ["SAM-2026-FL-001", "Food Supply — FCI Coleman Complex", "Federal Bureau of Prisons", "Federal",
         "RFQ", 250000, datetime(2026, 5, 15), '=INT(G2-TODAY())',
         82, "HIGH", "New",
         "Small Business Set-Aside", "Broad-line grocery, 3x/week delivery, cold chain required",
         "TBD — Food Service Administrator", "Not Started", "Largest federal prison complex in US. High volume."],
        ["SAM-2026-FL-002", "Fresh Produce Supply — VA Tampa", "Department of Veterans Affairs", "Federal",
         "RFQ", 120000, datetime(2026, 4, 20), '=INT(G3-TODAY())',
         75, "HIGH", "New",
         "Small Business Set-Aside", "Fresh fruits & vegetables, daily delivery, HACCP required",
         "NFS Manager — James A. Haley VAMC", "Not Started", "Major VA medical center. Weekly produce orders."],
        ["MFMP-2026-003", "School Nutrition Food Supply — Polk County", "Polk County School District", "Local",
         "IFB", 180000, datetime(2026, 6, 1), '=INT(G4-TODAY())',
         88, "HIGH", "New",
         "Local Preference", "Broad-line grocery + frozen items for 105K students",
         "Food Service Director", "Not Started", "Mid-size district. Less national competition."],
        ["MFMP-2026-004", "Jail Food Supply — Hillsborough County", "Hillsborough County Sheriff", "Local",
         "RFP", 320000, datetime(2026, 7, 15), '=INT(G5-TODAY())',
         70, "HIGH", "New",
         "None", "Inmate meal ingredients, 3 meals/day, ~3,500 inmates",
         "Purchasing Dept Contact", "Not Started", "Large county jail. Check if self-operated or contracted."],
        ["DS-2026-005", "Canned & Dry Goods — Brevard County Schools", "Brevard Public Schools", "Local",
         "RFQ", 85000, datetime(2026, 5, 30), '=INT(G6-TODAY())',
         65, "MEDIUM", "New",
         "Local Preference", "Canned vegetables, dry pasta, rice, beans — school cafeteria supply",
         "Food Service Director", "Not Started", "Space Coast district. 75K students."],
        ["SAM-2026-FL-006", "Dairy Products — NAS Jacksonville", "Defense Logistics Agency", "Federal",
         "Sources Sought", 200000, datetime(2026, 4, 10), '=INT(G7-TODAY())',
         55, "MEDIUM", "New",
         "Full & Open", "Milk, cheese, yogurt for base dining facilities",
         "DLA Troop Support", "Not Started", "Pre-solicitation. Respond to shape requirements."],
        ["MFMP-2026-007", "Snack & Beverage Supply — Seminole County Schools", "Seminole County Public Schools", "Local",
         "RFQ", 60000, datetime(2026, 6, 15), '=INT(G8-TODAY())',
         60, "MEDIUM", "New",
         "Local Preference", "Vending and a la carte snacks, beverages for 67K students",
         "Food Service Director", "Not Started", "Good candy/snack entry point."],
        ["SAM-2026-FL-008", "Emergency Food Staging — FEMA Region IV", "FEMA", "Federal",
         "Pre-Solicitation", 500000, datetime(2026, 8, 1), '=INT(G9-TODAY())',
         45, "MEDIUM", "New",
         "Small Business", "Pre-positioned food for hurricane season. FL warehouse required.",
         "FEMA Procurement", "Not Started", "Hurricane season prep. Newport's warehouse = key advantage."],
        ["DS-2026-009", "Meat & Poultry — Lee County Schools", "Lee County School District", "Local",
         "IFB", 95000, datetime(2026, 5, 1), '=INT(G10-TODAY())',
         50, "MEDIUM", "New",
         "Local Preference", "Ground beef, chicken, turkey for school meal program",
         "Food Service Director", "Not Started", "Fort Myers area. Growing district."],
        ["MFMP-2026-010", "Bakery Products — FL Dept of Corrections", "Florida DOC", "State",
         "RFQ", 150000, datetime(2026, 9, 1), '=INT(G11-TODAY())',
         35, "LOW", "New",
         "None", "Bread, rolls, baked goods for state prison system",
         "MFMP Category Manager", "Not Started", "State DOC uses Aramark. May be sub-supply opportunity."],
    ]

    for r, row_data in enumerate(sample, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 6:
                cell.number_format = CURRENCY_FMT
            elif c == 7 and isinstance(val, datetime):
                cell.number_format = DATE_FMT

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(headers), 2, len(sample) + 1)

    # Priority tier conditional formatting
    tier_col = "J"
    ws.conditional_formatting.add(
        f"{tier_col}2:{tier_col}100",
        CellIsRule(operator="equal", formula=['"HIGH"'], fill=HIGH_FILL),
    )
    ws.conditional_formatting.add(
        f"{tier_col}2:{tier_col}100",
        CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=MED_FILL),
    )
    ws.conditional_formatting.add(
        f"{tier_col}2:{tier_col}100",
        CellIsRule(operator="equal", formula=['"LOW"'], fill=LOW_FILL),
    )

    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


# --- Tab 2: Expiring Contracts ---
def build_expiring_tab(wb):
    ws = wb.create_sheet("Expiring Contracts")
    headers = [
        "Contract Number", "Current Vendor", "Incumbent Performance",
        "Agency", "Facility", "Annual Value",
        "Expiration Date", "Months Until Expiry",
        "Contracting Officer", "Food Service Admin",
        "Entry Opportunity Score", "Outreach Status", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    sample = [
        ["BOP-FL-2022-0145", "Sysco Government", "Good",
         "Bureau of Prisons", "FCI Coleman Complex", 450000,
         datetime(2027, 3, 31), '=INT((G2-TODAY())/30)',
         "TBD — FPDS Lookup", "TBD — Call Institution",
         65, "Not Started", "Largest BOP complex. High value. Strong incumbent."],
        ["VA-FL-2023-0089", "US Foods", "Mixed",
         "VA — James A. Haley", "Tampa VAMC", 180000,
         datetime(2026, 9, 30), '=INT((G3-TODAY())/30)',
         "TBD — FPDS Lookup", "NFS Manager",
         78, "Not Started", "Mixed reviews on delivery consistency. Opportunity."],
        ["PBCSD-2023-FN-042", "Local Distributor", "Unknown",
         "Palm Beach County Schools", "District-wide", 350000,
         datetime(2026, 12, 31), '=INT((G4-TODAY())/30)',
         "Procurement Dept", "Food Service Director",
         82, "Not Started", "195K students. Contract expiring. Strong entry point."],
        ["HCSO-2024-FOOD-001", "Aramark", "Poor",
         "Hillsborough County Sheriff", "Orient Road Jail", 280000,
         datetime(2027, 6, 30), '=INT((G5-TODAY())/30)',
         "County Purchasing", "Jail Food Service Mgr",
         72, "Not Started", "Aramark performance complaints. Tampa area."],
        ["USDA-FL-2023-AMS-017", "Regional Distributor", "Good",
         "USDA AMS", "FL Commodity Distribution", 200000,
         datetime(2026, 8, 31), '=INT((G6-TODAY())/30)',
         "AMS Commodity Procurement", "State USDA Contact",
         58, "Not Started", "USDA commodity program. Requires QBL approval."],
    ]

    for r, row_data in enumerate(sample, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 6:
                cell.number_format = CURRENCY_FMT
            elif c == 7 and isinstance(val, datetime):
                cell.number_format = DATE_FMT

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(headers), 2, len(sample) + 1)
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


# --- Tab 3: Sources Sought Tracker ---
def build_sources_sought_tab(wb):
    ws = wb.create_sheet("Sources Sought Tracker")
    headers = [
        "Notice ID", "Title", "Agency", "Posted Date",
        "Response Deadline", "Days to Respond",
        "Response Status", "Response Submitted Date",
        "Follow-Up Actions", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    sample = [
        ["SS-2026-BOP-FL-001", "Market Research: Food Supply Services — SE Region",
         "Bureau of Prisons", datetime(2026, 2, 10), datetime(2026, 3, 10),
         '=INT(E2-TODAY())', "Drafting", "",
         "Request Industry Day meeting", "BOP exploring new vendor options for SE institutions"],
        ["SS-2026-USDA-003", "Sources Sought: Fresh Produce Distribution — FL",
         "USDA AMS", datetime(2026, 2, 15), datetime(2026, 3, 15),
         '=INT(E3-TODAY())', "Not Started", "",
         "Submit capability overview", "USDA expanding FL commodity distribution network"],
        ["SS-2026-FEMA-012", "Pre-Disaster Food Staging — FEMA Region IV",
         "FEMA", datetime(2026, 1, 20), datetime(2026, 3, 1),
         '=INT(E4-TODAY())', "Not Started", "",
         "Highlight FL warehouse capacity", "Hurricane season preparedness. Newport's warehouse is the pitch."],
    ]

    for r, row_data in enumerate(sample, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if isinstance(val, datetime):
                cell.number_format = DATE_FMT

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(headers), 2, len(sample) + 1)
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


# --- Tab 4: Decision Maker Tracker ---
def build_contacts_tab(wb, contacts_csv=None):
    ws = wb.create_sheet("Decision Makers")
    headers = [
        "Name", "Title", "Organization", "Organization Type",
        "Role Type", "Tier", "Influence Level", "Geography",
        "Source", "Email Available", "Phone Available",
        "Contact Status", "Last Contact Date", "Next Action",
        "Current Vendor", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    if contacts_csv and Path(contacts_csv).exists():
        with open(contacts_csv) as f:
            reader = csv.DictReader(f)
            for r, row in enumerate(reader, 2):
                ws.cell(row=r, column=1, value=row.get("name", ""))
                ws.cell(row=r, column=2, value=row.get("title", ""))
                ws.cell(row=r, column=3, value=row.get("organization", ""))
                ws.cell(row=r, column=4, value=row.get("org_type", ""))
                ws.cell(row=r, column=5, value=row.get("role_type", ""))
                ws.cell(row=r, column=6, value=row.get("tier", ""))
                ws.cell(row=r, column=7, value=row.get("influence", ""))
                ws.cell(row=r, column=8, value=row.get("geography", ""))
                ws.cell(row=r, column=9, value=row.get("source", ""))
                ws.cell(row=r, column=10, value=row.get("email_available", ""))
                ws.cell(row=r, column=11, value=row.get("phone_available", ""))
                ws.cell(row=r, column=12, value=row.get("contact_status", "Not Started"))
                ws.cell(row=r, column=13, value=row.get("last_contact", ""))
                ws.cell(row=r, column=14, value=row.get("next_action", ""))
                ws.cell(row=r, column=15, value=row.get("current_vendor", ""))
                ws.cell(row=r, column=16, value=row.get("notes", ""))
    else:
        # Placeholder rows showing the structure
        placeholder = [
            ["[Food Service Director]", "Director of Food & Nutrition Services",
             "Miami-Dade County Public Schools", "Local — School District",
             "Food Service", "Tier 1", "High", "Miami-Dade County, FL",
             "District Website", "Yes", "Yes",
             "Not Started", "", "Schedule introductory meeting",
             "Unknown", "Largest FL district. 350K+ students. Priority target."],
            ["[Food Service Administrator]", "Food Service Administrator",
             "FCI Coleman Complex", "Federal — BOP",
             "Food Service", "Tier 1", "High", "Sumter County, FL",
             "BOP.gov / Phone", "Unknown", "Yes (institution main line)",
             "Not Started", "", "Call institution, request meeting",
             "Sysco Government", "Largest federal prison complex in US. 8,000+ inmates."],
            ["[Contracting Officer]", "Contracting Officer",
             "VA — James A. Haley VAMC", "Federal — VA",
             "Contracting", "Tier 2", "High", "Tampa, FL",
             "FPDS Award Data", "Yes (.va.gov format)", "Yes",
             "Not Started", "", "Submit capability statement",
             "US Foods", "Tampa VA medical center. Signed recent food contracts."],
        ]
        for r, row_data in enumerate(placeholder, 2):
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val)

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


# --- Tab 5: Past Performance Portfolio ---
def build_past_performance_tab(wb):
    ws = wb.create_sheet("Past Performance")
    headers = [
        "Contract/PO Number", "Agency/Client", "Description",
        "Value", "Period Start", "Period End",
        "On-Time Delivery Rate", "Quality Issues",
        "Reference Letter", "Reference Contact", "CPARS Rating", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    ws.cell(row=2, column=1, value="[To be populated as contracts are won]")
    ws.cell(row=3, column=1, value="")
    ws.cell(row=4, column=1, value="INSTRUCTIONS:")
    ws.cell(row=5, column=1, value="1. After every government delivery, log it here immediately")
    ws.cell(row=6, column=1, value="2. Request a written reference letter after every successful delivery")
    ws.cell(row=7, column=1, value="3. Track on-time rate and quality issues — this data goes into future proposals")
    ws.cell(row=8, column=1, value="4. For federal contracts, monitor CPARS ratings")
    ws.cell(row=9, column=1, value="5. Goal: 5+ documented references in Year 1")

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


# --- Tab 6: Scoring Model ---
def build_scoring_tab(wb):
    ws = wb.create_sheet("Scoring Model")

    title_font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    subtitle_font = Font(name="Calibri", bold=True, size=11, color=NAVY)

    ws.cell(row=1, column=1, value="Opportunity Priority Scoring Model").font = title_font
    ws.cell(row=2, column=1, value="Score range: 0-100. Higher = higher priority for Newport.").font = DATA_FONT

    headers = ["Factor", "Points", "HIGH Score Criteria", "MEDIUM Score Criteria", "LOW Score Criteria"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, len(headers), row=4)

    model = [
        ["Set-Aside Status", 25,
         "Small Business set-aside (25 pts)", "HUBZone or other preference (15 pts)", "Full & Open (5 pts)"],
        ["Past Performance Requirement", 20,
         "None required (20 pts)", "Minimal / private-sector OK (15 pts)", "Significant gov experience needed (5 pts)"],
        ["Geographic Fit", 20,
         "Florida — within delivery range (20 pts)", "Southeast US (10 pts)", "National / out of range (0 pts)"],
        ["Contract Value", 15,
         "$100K+ annual value (15 pts)", "$25K-$100K (10 pts)", "Under $25K (5 pts)"],
        ["Competition Level", 10,
         "Few bidders expected / sole source (10 pts)", "Moderate competition (5 pts)", "Heavy competition (0 pts)"],
        ["Relationship Status", 10,
         "Active relationship with decision maker (10 pts)", "Some prior contact (5 pts)", "No relationship (0 pts)"],
    ]

    for r, row_data in enumerate(model, 5):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, len(headers), 5, 10)

    ws.cell(row=12, column=1, value="Priority Tiers").font = subtitle_font
    tiers = [
        ["HIGH (70-100)", "Actively pursue — assign to salesperson immediately", HIGH_FILL],
        ["MEDIUM (40-69)", "Evaluate — worth a response if bandwidth allows", MED_FILL],
        ["LOW (0-39)", "Monitor — track for future reference, no action now", LOW_FILL],
    ]
    for r, (tier, desc, fill) in enumerate(tiers, 13):
        cell_tier = ws.cell(row=r, column=1, value=tier)
        cell_tier.fill = fill
        cell_tier.font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=r, column=2, value=desc).font = DATA_FONT

    auto_width(ws, len(headers), max_width=50)
    return ws


# --- Tab 7: Summary Dashboard ---
def build_summary_tab(wb, market_data_csv=None):
    ws = wb.create_sheet("Summary Dashboard")

    title_font = Font(name="Calibri", bold=True, size=16, color=NAVY)
    section_font = Font(name="Calibri", bold=True, size=13, color=NAVY)
    kpi_value_font = Font(name="Calibri", bold=True, size=24, color=NAVY)
    kpi_label_font = Font(name="Calibri", size=10, color="666666")
    gold_fill = PatternFill(start_color="FFF8E7", end_color="FFF8E7", fill_type="solid")

    ws.cell(row=1, column=1, value="Newport Wholesalers — Government Contract Intelligence Dashboard").font = title_font
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y')}").font = kpi_label_font

    # KPI Cards
    ws.cell(row=4, column=1, value="MARKET OPPORTUNITY").font = section_font

    kpis = [
        ("A", 6, "$35-50B+", "Total US Gov Food Market (Annual)"),
        ("C", 6, "$1.47B", "Federal Food Spend FY2025 (Tracked)"),
        ("E", 6, "$500M-$2B", "FL Serviceable Market (Annual)"),
        ("G", 6, "15%", "Federal Spend Growth (2-Year)"),
    ]
    for col_letter, row, value, label in kpis:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell_val = ws.cell(row=row, column=col_idx, value=value)
        cell_val.font = kpi_value_font
        cell_val.fill = gold_fill
        cell_val.alignment = Alignment(horizontal="center")
        cell_label = ws.cell(row=row + 1, column=col_idx, value=label)
        cell_label.font = kpi_label_font
        cell_label.alignment = Alignment(horizontal="center")

    # Federal Spending by Agency
    ws.cell(row=9, column=1, value="FEDERAL FOOD SPENDING BY AGENCY (FY2025)").font = section_font

    agency_headers = ["Agency", "FY2023", "FY2024", "FY2025", "2-Year Growth"]
    for col, h in enumerate(agency_headers, 1):
        ws.cell(row=10, column=col, value=h)
    style_header_row(ws, len(agency_headers), row=10)

    agencies = []
    if market_data_csv and Path(market_data_csv).exists():
        by_year = {}
        with open(market_data_csv) as f:
            reader = csv.DictReader(f)
            for row in reader:
                agency = row["agency"]
                fy = row["fiscal_year"]
                amount = float(row["amount"])
                if agency not in by_year:
                    by_year[agency] = {}
                by_year[agency][fy] = amount
        for agency, years in sorted(by_year.items(), key=lambda x: x[1].get("FY2025", 0), reverse=True)[:8]:
            fy23 = years.get("FY2023", 0)
            fy24 = years.get("FY2024", 0)
            fy25 = years.get("FY2025", 0)
            growth = ((fy25 / fy23) - 1) * 100 if fy23 > 0 else 0
            agencies.append([agency, fy23, fy24, fy25, growth])
    else:
        agencies = [
            ["Department of Defense", 1085033880, 1105371602, 1193217505, 10.0],
            ["Department of Agriculture", 66026700, 81988223, 117089691, 77.3],
            ["Department of Homeland Security", 73179075, 79876738, 90170437, 23.2],
            ["Department of Justice (BOP)", 16568594, 15887863, 19742994, 19.1],
            ["Department of State", 7035660, 7034532, 12354730, 75.6],
            ["Department of the Treasury", 8186221, 7917739, 10071552, 23.0],
            ["Department of Veterans Affairs", 9193022, 8013431, 9044673, -1.6],
            ["Department of Transportation", 6040275, 8112314, 6083904, 0.7],
        ]

    for r, row_data in enumerate(agencies, 11):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (2, 3, 4):
                cell.number_format = CURRENCY_FMT
            elif c == 5:
                cell.number_format = '0.0"%"'

    style_data_rows(ws, len(agency_headers), 11, 11 + len(agencies) - 1)

    # Fraud environment
    fraud_row = 11 + len(agencies) + 2
    ws.cell(row=fraud_row, column=1, value="WHY NOW — THE FRAUD CRISIS CREATES OPPORTUNITY").font = section_font

    fraud_kpis = [
        ("A", fraud_row + 2, "$6.8B", "FCA Recoveries FY2025 (Record)"),
        ("C", fraud_row + 2, "1,091", "8(a) Firms Suspended (25%)"),
        ("E", fraud_row + 2, "$250M+", "Feeding Our Future Fraud"),
        ("G", fraud_row + 2, "10,000+", "DOGE Contracts Terminated"),
    ]
    for col_letter, row, value, label in fraud_kpis:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell_val = ws.cell(row=row, column=col_idx, value=value)
        cell_val.font = Font(name="Calibri", bold=True, size=20, color="CC0000")
        cell_val.alignment = Alignment(horizontal="center")
        cell_label = ws.cell(row=row + 1, column=col_idx, value=label)
        cell_label.font = kpi_label_font
        cell_label.alignment = Alignment(horizontal="center")

    # Newport's edge
    edge_row = fraud_row + 5
    ws.cell(row=edge_row, column=1, value="NEWPORT'S COMPETITIVE EDGE").font = section_font
    edges = [
        "25+ years of continuous Florida operations — provable, auditable, real",
        "American-owned, American employees, American taxes",
        "Real infrastructure: warehouses, trucks, cold chain",
        "Small Business qualification under NAICS 424410 ($200M threshold) — locks out Sysco, US Foods, Aramark",
        "Florida home state advantage: local preference scoring on FL contracts",
        "FEMA disaster staging: FL warehouse + hurricane season = sole-source opportunity",
    ]
    for i, edge in enumerate(edges):
        ws.cell(row=edge_row + 1 + i, column=1, value=f"  {edge}").font = DATA_FONT

    # Pipeline summary
    pipeline_row = edge_row + len(edges) + 3
    ws.cell(row=pipeline_row, column=1, value="PIPELINE SNAPSHOT").font = section_font
    pipeline_data = [
        ["Active Opportunities", "10", "See 'Opportunities Pipeline' tab"],
        ["Expiring Contracts Tracked", "5", "See 'Expiring Contracts' tab"],
        ["Sources Sought to Respond", "3", "See 'Sources Sought Tracker' tab"],
        ["Decision Makers Identified", "68+", "See 'Decision Makers' tab"],
        ["Past Performance References", "0 (building)", "See 'Past Performance' tab"],
    ]
    for r, (metric, value, note) in enumerate(pipeline_data, pipeline_row + 1):
        ws.cell(row=r, column=1, value=metric).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=r, column=2, value=value).font = Font(name="Calibri", bold=True, size=12, color=NAVY)
        ws.cell(row=r, column=3, value=note).font = kpi_label_font

    # Set column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 18

    ws.sheet_properties.tabColor = GOLD
    return ws


# --- Tab 8: Conference & Event Calendar ---
def build_events_tab(wb):
    ws = wb.create_sheet("Events Calendar")
    headers = [
        "Event", "Association", "Typical Date", "Location",
        "Target Contacts", "Registration Cost", "Registration Status",
        "Newport Attendee", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    events = [
        ["FL School Nutrition Association Conference", "FSNA / SNA", "Spring 2026 (TBD)", "Florida",
         "FL school district food service directors", "$75 membership + $200-400 conference", "Not Registered",
         "TBD", "PRIORITY: Every FL school nutrition director in one room"],
        ["ACFSA Annual Conference", "American Correctional Food Service Association", "August 2026", "Varies (national)",
         "Corrections food service administrators nationwide", "$150 membership + $500-800 conference", "Not Registered",
         "TBD", "Best networking for prison food service contacts"],
        ["ALA Annual Conference & EXPO", "American Logistics Association", "January 2027", "Washington DC area",
         "Military commissary buyers, DLA personnel, DeCA contacts", "$500 membership + $800-1200 conference", "Not Registered",
         "TBD", "Military food supply chain. Premium event."],
        ["FL Association of Counties Annual Conference", "FAC", "June 2026", "Florida",
         "County purchasing officials across all 67 FL counties", "Varies", "Not Registered",
         "TBD", "Access to county jail and facility food purchasing contacts"],
        ["School Nutrition Association ANC", "SNA (National)", "July 2026", "Varies (national)",
         "10,000+ school nutrition professionals nationwide", "$75 membership + $500-700 conference", "Not Registered",
         "TBD", "National event. Attend after establishing FL presence."],
        ["Florida Restaurant & Lodging Show", "FRLA", "Fall 2026", "Orlando, FL",
         "FL food service industry contacts", "Varies", "Not Registered",
         "TBD", "Secondary priority. Broader food industry networking."],
    ]

    for r, row_data in enumerate(events, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(headers), 2, len(events) + 1)
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


# --- Main ---
def main():
    parser = argparse.ArgumentParser(description="Generate Newport opportunity dashboard")
    parser.add_argument("--contacts", help="Path to contacts CSV for Decision Makers tab")
    parser.add_argument("--market-data", help="Path to market data CSV for Summary tab")
    parser.add_argument("--output", help="Output file path", default=None)
    args = parser.parse_args()

    market_csv = args.market_data
    if not market_csv:
        # Try to find market data automatically
        candidates = sorted(PROJECT_ROOT.glob("data/final/govt_market_size_by_agency_*.csv"))
        if candidates:
            market_csv = str(candidates[-1])
            print(f"  Auto-detected market data: {market_csv}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building dashboard tabs...")
    print("  Tab 7: Summary Dashboard")
    build_summary_tab(wb, market_csv)
    print("  Tab 1: Opportunities Pipeline")
    build_opportunities_tab(wb)
    print("  Tab 2: Expiring Contracts")
    build_expiring_tab(wb)
    print("  Tab 3: Sources Sought Tracker")
    build_sources_sought_tab(wb)
    print("  Tab 4: Decision Makers")
    build_contacts_tab(wb, args.contacts)
    print("  Tab 5: Past Performance")
    build_past_performance_tab(wb)
    print("  Tab 6: Scoring Model")
    build_scoring_tab(wb)
    print("  Tab 8: Events Calendar")
    build_events_tab(wb)

    output_path = args.output or str(
        OUTPUT_DIR / f"newport_dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    wb.save(output_path)
    print(f"\nDashboard saved to: {output_path}")
    print(f"  {wb.sheetnames}")


if __name__ == "__main__":
    main()
